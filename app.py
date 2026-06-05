import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import hashlib
from dataclasses import dataclass, field
from datetime import datetime, date
import time
import json
import logging
import difflib

import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─── Optimization Module ───
try:
    from streamlit_optimizations import (
        optimize_dtypes,
        create_lazy_preview,
        create_safe_styler,
        safe_dataframe_display,
        display_data_summary,
        init_session_state_optimized,
    )
except ImportError:
    logger.warning("Optimization module streamlit_optimizations not available; using built-in fallback.")
    # Fallback functions
    def optimize_dtypes(df, verbose=False):
        return df, {}
    def create_lazy_preview(df, preview_rows=100):
        return df.head(preview_rows), {"total_rows": len(df), "total_cols": len(df.columns)}
    def create_safe_styler(df, *args, **kwargs):
        return df.style
    def safe_dataframe_display(df, *args, **kwargs):
        st.dataframe(df, use_container_width=True)
    def display_data_summary(df, *args, **kwargs):
        pass
    def init_session_state_optimized():
        return None

# ─── Advanced Features ───
try:
    from core.text_processor import (
        AdvancedCleaner,
        AuditTrail,
        clean_phone_column,
        detect_phone_column,
        clean_numeric_columns,
        detect_numeric_columns_with_text,
        apply_fuzzy_matching,
        optimize_memory,
        read_csv_with_encoding_fallback,
    )
    from core.date_cleaner import clean_dates_advanced
    from stats.advanced_imputation import (
        universal_missing_value_imputation,
        streamlit_imputation_style,
    )
except ImportError:
    st.warning("⚠️ Advanced cleaning features not available. Install requirements: pip install -r requirements.txt")
    AdvancedCleaner = None

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="DataClean Pro | بوابة تنظيف البيانات",
    page_icon="✦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── FIX: Disable verbose client error reporting & handle DOM errors ───
st.markdown("""
<script>
    // Suppress removeChild errors gracefully
    (function() {
        const originalRemoveChild = Node.prototype.removeChild;
        Node.prototype.removeChild = function(child) {
            try {
                return originalRemoveChild.call(this, child);
            } catch (e) {
                if (e.name === 'NotFoundError' && e.message.includes('removeChild')) {
                    console.warn('Safe removeChild error handled:', e);
                    return child;
                }
                throw e;
            }
        };
        window.streamlitShowErrorDetails = false;
    })();
</script>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# DESIGN SYSTEM
# ═══════════════════════════════════════════════════════════════════════════════
C = {
    "bg":        "#070B14",
    "bg1":       "#0D1117",
    "bg2":       "#111827",
    "bg3":       "#1A2332",
    "border":    "#1E2D3D",
    "border2":   "#2D3748",
    "accent":    "#3B82F6",
    "accent2":   "#6366F1",
    "accent3":   "#8B5CF6",
    "cyan":      "#06B6D4",
    "green":     "#10B981",
    "yellow":    "#F59E0B",
    "red":       "#EF4444",
    "text":      "#FFFFFF",
    "text2":     "#94A3B8",
    "text3":     "#64748B",
    "white":     "#FFFFFF",
}

# ═══════════════════════════════════════════════════════════════════════════════
# MISSING VALUES — ARABIC + ENGLISH
# ═══════════════════════════════════════════════════════════════════════════════
MISSING_VALUES = {
    # English
    "", "unknown", "na", "n/a", "none", "null", "nan", "nat",
    "#n/a", "-", "--", "?", "missing", "not available",
    "undefined", "not specified", "not provided", "not found",
    "not applicable", "no data", "no value", "empty",
    "n.a", "n.a.", "nil", "void", "blank",
    # Arabic
    "غير معروف", "غير معرف", "غير محدد", "لا يوجد",
    "مفقود", "غير متاح", "غير موجود", "غير مدرج",
    "غير معرّف", "لا شيء", "فارغ", "غير مكتمل",
}

# ═══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
# CSS
# ═══════════════════════════════════════════════════════════════════════════════
def inject_css():
    st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@300;400;500;700;800;900&family=Inter:wght@300;400;500;600;700;800&display=swap');

    /* ─── FIX: Prevent animation-related DOM errors ─── */
    * {{
        box-sizing: border-box;
        will-change: auto !important;
    }}
    
    *::before, *::after {{ box-sizing: border-box; }}

    html, body,
    [data-testid="stAppViewContainer"],
    [data-testid="stMain"],
    .main, .block-container {{
        background-color: {C['bg']} !important;
        color: {C['text']};
        font-family: 'Tajawal', 'Inter', sans-serif !important;
        direction: rtl;
    }}
    .block-container {{
        padding: 2rem 3rem !important;
        max-width: 1400px !important;
    }}
    [data-testid="stSidebar"] {{
        background: {C['bg1']} !important;
        border-left: 1px solid {C['border']} !important;
    }}
    [data-testid="stSidebar"] * {{
        font-family: 'Tajawal', 'Inter', sans-serif !important;
        background: transparent !important;
    }}
    section[data-testid="stSidebar"] {{ width: 290px !important; max-width: 100% !important; }}

    h1,h2,h3,h4,h5,h6 {{
        font-family: 'Tajawal', 'Inter', sans-serif !important;
        color: {C['text']} !important;
        font-weight: 800 !important;
    }}

    /* English text inside Arabic layout */
    .en-text {{
        font-family: 'Inter', sans-serif !important;
        direction: ltr;
        display: inline-block;
    }}
    .mixed-cell {{
        font-family: 'Tajawal', 'Inter', sans-serif !important;
    }}

    ::-webkit-scrollbar {{ width:6px; height:6px; }}
    ::-webkit-scrollbar-track {{ background:{C['bg1']}; }}
    ::-webkit-scrollbar-thumb {{ background:{C['border2']}; border-radius:3px; }}
    ::-webkit-scrollbar-thumb:hover {{ background:{C['accent']}; }}

    .stButton > button {{
        font-family: 'Tajawal', 'Inter', sans-serif !important;
        font-weight: 700 !important;
        font-size: 0.97rem !important;
        border-radius: 12px !important;
        border: none !important;
        padding: 0.8rem 1.8rem !important;
        width: 100% !important;
        transition: all 0.3s ease !important;
        background: linear-gradient(135deg, {C['accent']}, {C['accent2']}) !important;
        color: {C['white']} !important;
        box-shadow: 0 4px 15px rgba(59,130,246,0.25) !important;
    }}
    .stButton > button:hover {{
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 25px rgba(59,130,246,0.4) !important;
    }}
    .stButton > button:active {{
        transform: translateY(0px) !important;
    }}
    .stDownloadButton > button {{
        font-family: 'Tajawal', 'Inter', sans-serif !important;
        font-weight: 700 !important;
        font-size: 0.97rem !important;
        border-radius: 12px !important;
        border: none !important;
        padding: 0.8rem 1.8rem !important;
        width: 100% !important;
        transition: all 0.3s ease !important;
        background: linear-gradient(135deg, {C['green']}, #059669) !important;
        color: {C['white']} !important;
        box-shadow: 0 4px 15px rgba(16,185,129,0.25) !important;
    }}
    .stDownloadButton > button:hover {{
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 25px rgba(16,185,129,0.4) !important;
    }}
    [data-testid="stFileUploader"] {{
        background: linear-gradient(135deg, {C['bg2']}, {C['bg1']}) !important;
        border: 2px dashed {C['border2']} !important;
        border-radius: 16px !important;
        padding: 2.4rem !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1) !important;
    }}
    [data-testid="stFileUploader"]:hover {{
        border-color: {C['accent']} !important;
        background: linear-gradient(135deg, rgba(59,130,246,0.08), rgba(99,102,241,0.04)) !important;
        box-shadow: 0 8px 20px rgba(59,130,246,0.15) !important;
    }}
    [data-testid="stFileUploader"] label,
    [data-testid="stFileUploader"] p,
    [data-testid="stFileUploader"] span {{
        color: {C['text2']} !important;
        font-family: 'Tajawal', 'Inter', sans-serif !important;
    }}
    [data-testid="stFileUploader"] button {{
        background: linear-gradient(135deg, {C['accent']}, {C['accent2']}) !important;
        border: none !important;
        color: {C['text']} !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        transition: all 0.3s ease !important;
    }}
    [data-testid="stFileUploader"] button:hover {{
        transform: translateY(-2px) !important;
        box-shadow: 0 4px 12px rgba(59,130,246,0.25) !important;
    }}
    [data-testid="stFileUploader"] button [data-testid="stIconMaterial"] {{
        display: none !important;
    }}
    .stTabs [data-baseweb="tab-list"] {{
        background: linear-gradient(135deg, {C['bg2']}, {C['bg1']}) !important;
        border-radius: 14px !important;
        padding: 0.5rem !important;
        gap: 0.5rem !important;
        border: 1px solid {C['border']} !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1) !important;
    }}
    .stTabs [data-baseweb="tab"] {{
        border-radius: 10px !important;
        font-family: 'Tajawal', 'Inter', sans-serif !important;
        font-weight: 600 !important;
        color: {C['text2']} !important;
        padding: 0.65rem 1.4rem !important;
        transition: all 0.3s ease !important;
        border: none !important;
        background: transparent !important;
    }}
    .stTabs [aria-selected="true"] {{
        background: linear-gradient(135deg, {C['accent']}, {C['accent2']}) !important;
        color: {C['white']} !important;
        box-shadow: 0 4px 12px rgba(59,130,246,0.3) !important;
    }}
    .stTabs [data-baseweb="tab-panel"] {{
        background: transparent !important;
        padding: 1.5rem 0 0 !important;
    }}
    .stExpander {{
        background: linear-gradient(135deg, {C['bg2']}, {C['bg1']}) !important;
        border: 1px solid {C['border']} !important;
        border-radius: 14px !important;
        overflow: hidden !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05) !important;
    }}
    .stExpander:hover {{
        border-color: {C['border2']} !important;
    }}
    div[data-testid="stRadio"] > div {{ gap: 1rem !important; }}
    div[data-testid="stRadio"] label {{
        background: linear-gradient(135deg, {C['bg2']}, {C['bg1']}) !important;
        border: 1.5px solid {C['border2']} !important;
        border-radius: 12px !important;
        padding: 0.8rem 1.5rem !important;
        transition: all 0.25s ease !important;
        cursor: pointer !important;
        color: {C['text']} !important;
        font-family: 'Tajawal', 'Inter', sans-serif !important;
        font-weight: 600 !important;
        box-shadow: 0 2px 6px rgba(0,0,0,0.05) !important;
    }}
    div[data-testid="stRadio"] label:hover {{
        border-color: {C['accent']} !important;
        background: linear-gradient(135deg, rgba(59,130,246,0.08), rgba(99,102,241,0.04)) !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 4px 12px rgba(59,130,246,0.15) !important;
    }}
    [data-testid="stDataFrame"] {{
        border: 1px solid {C['border']} !important;
        border-radius: 14px !important;
        overflow: hidden !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05) !important;
    }}
    [data-testid="stProgress"] > div > div {{
        background: linear-gradient(90deg, {C['accent']}, {C['accent2']}) !important;
        border-radius: 10px !important;
        box-shadow: 0 0 12px rgba(59,130,246,0.3) !important;
    }}
    [data-testid="stProgress"] > div {{
        background: {C['bg3']} !important;
        border-radius: 10px !important;
    }}
    hr {{
        border: none !important;
        border-top: 1px solid {C['border']} !important;
        margin: 1.5rem 0 !important;
    }}

    /* ══ CUSTOM COMPONENTS ══ */
    .hero {{
        text-align: center;
        padding: 3rem 2rem 2rem;
        background: linear-gradient(135deg, rgba(59,130,246,0.05) 0%, rgba(99,102,241,0.05) 100%);
        border-radius: 20px;
        border: 1px solid {C['border']};
        margin-bottom: 1.5rem;
        position: relative;
        overflow: hidden;
    }}
    .hero::before {{
        content: '';
        position: absolute;
        top: -50%; right: -50%;
        width: 300px; height: 300px;
        background: radial-gradient(circle, {C['accent']}08 0%, transparent 70%);
        border-radius: 50%;
        z-index: 0;
    }}
    .hero {{
        position: relative;
        z-index: 1;
    }}
    .hero-badge {{
        display: inline-flex;
        align-items: center;
        gap: 0.6rem;
        background: linear-gradient(135deg, rgba(59,130,246,0.12), rgba(59,130,246,0.06));
        border: 1px solid rgba(59,130,246,0.35);
        border-radius: 50px;
        padding: 0.5rem 1.4rem;
        font-size: 0.85rem;
        font-weight: 700;
        color: {C['accent']};
        margin-bottom: 1.2rem;
        letter-spacing: 0.06em;
        text-transform: uppercase;
        box-shadow: inset 0 1px 0 rgba(255,255,255,0.1);
    }}
    .hero-title {{
        font-size: 3rem;
        font-weight: 900;
        line-height: 1.15;
        margin-bottom: 1rem;
        color: {C['text']};
        letter-spacing: -0.02em;
    }}
    .hero-title span {{
        background: linear-gradient(135deg, {C['accent']}, {C['cyan']});
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        filter: drop-shadow(0 0 10px rgba(59,130,246,0.15));
    }}
    .hero-subtitle {{
        font-size: 1.05rem;
        color: {C['text2']};
        max-width: 620px;
        margin: 0 auto 1.5rem;
        line-height: 1.8;
        font-weight: 500;
    }}
    .lang-badge {{
        display: inline-flex;
        align-items: center;
        gap: 0.5rem;
        background: linear-gradient(135deg, rgba(16,185,129,0.12), rgba(16,185,129,0.06));
        border: 1px solid rgba(16,185,129,0.3);
        border-radius: 50px;
        padding: 0.35rem 1.1rem;
        font-size: 0.82rem;
        font-weight: 700;
        color: {C['green']};
        margin: 0 0.4rem;
        text-transform: uppercase;
        letter-spacing: 0.04em;
    }}

    .stepper-wrap {{
        display: flex;
        justify-content: center;
        align-items: center;
        padding: 1.5rem 2rem;
        background: linear-gradient(135deg, {C['bg1']}, {C['bg2']});
        border: 1px solid {C['border']};
        border-radius: 16px;
        margin: 1.5rem 0;
        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    }}
    .stepper-step {{
        display: flex;
        flex-direction: column;
        align-items: center;
        gap: 0.5rem;
        min-width: 100px;
    }}
    .stepper-circle {{
        width: 50px; height: 50px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 800;
        font-size: 1rem;
        transition: all 0.3s ease;
        background: linear-gradient(135deg, {C['bg3']}, {C['bg2']});
        border: 2px solid {C['border']};
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }}
    .stepper-circle.active {{
        background: linear-gradient(135deg, {C['accent']}, {C['accent2']});
        border-color: {C['accent']};
        transform: scale(1.1);
        box-shadow: 0 4px 15px rgba(59,130,246,0.3);
    }}
    .stepper-label {{
        font-size: 0.85rem;
        font-weight: 700;
        text-align: center;
        color: {C['text2']};
    }}
    .stepper-line {{
        flex: 1; height: 2px;
        max-width: 70px; min-width: 30px;
        margin-bottom: 22px;
        border-radius: 2px;
        background: {C['border']};
        transition: all 0.3s ease;
    }}
    .stepper-line.active {{
        background: linear-gradient(90deg, {C['accent']}, {C['accent2']});
        box-shadow: 0 0 8px rgba(59,130,246,0.3);
    }}

    .metric-grid {{
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
        gap: 1.2rem;
        margin: 1.5rem 0;
    }}
    .metric-card {{
        background: linear-gradient(135deg, {C['bg2']}, {C['bg1']});
        border: 1px solid {C['border']};
        border-radius: 16px;
        padding: 1.4rem 1.2rem;
        text-align: center;
        transition: all 0.3s ease;
        position: relative;
        overflow: hidden;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
    }}
    .metric-card::before {{
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0; bottom: 0;
        background: linear-gradient(135deg, rgba(59,130,246,0.05), transparent);
        opacity: 0;
        transition: opacity 0.3s ease;
        pointer-events: none;
    }}
    .metric-card::after {{
        content: '';
        position: absolute;
        bottom: 0; left: 0; right: 0;
        height: 3px;
        background: linear-gradient(90deg, {C['accent']}, {C['accent2']});
        transform: scaleX(0);
        transition: transform 0.3s ease;
        transform-origin: left;
    }}
    .metric-card:hover {{
        border-color: {C['accent']}60;
        transform: translateY(-4px);
        box-shadow: 0 8px 24px rgba(59,130,246,0.2);
    }}
    .metric-card:hover::before {{ opacity: 1; }}
    .metric-card:hover::after {{ transform: scaleX(1); }}
    .metric-icon {{ font-size: 1.8rem; margin-bottom: 0.5rem; }}
    .metric-label {{
        font-size: 0.75rem;
        color: {C['text3']};
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.06em;
        margin-bottom: 0.4rem;
    }}
    .metric-value {{
        font-size: 2rem;
        font-weight: 900;
        color: {C['accent']};
        line-height: 1;
    }}
    .metric-sub {{ font-size: 0.72rem; color: {C['text3']}; margin-top: 0.3rem; }}

    .stage-header {{
        background: linear-gradient(135deg, {C['bg2']}, {C['bg1']});
        border: 1px solid {C['border']};
        border-radius: 18px;
        padding: 2rem 2.2rem;
        margin-bottom: 2rem;
        position: relative;
        overflow: hidden;
        box-shadow: 0 8px 24px rgba(0,0,0,0.15);
    }}
    .stage-header::before {{
        content: '';
        position: absolute;
        top: 0; right: 0;
        width: 250px; height: 250px;
        background: radial-gradient(circle, {C['accent']}10 0%, transparent 70%);
        border-radius: 50%;
        z-index: 0;
    }}
    .stage-header-num {{
        display: inline-flex;
        align-items: center;
        gap: 0.5rem;
        font-size: 0.78rem;
        font-weight: 700;
        color: {C['accent']};
        text-transform: uppercase;
        letter-spacing: 0.12em;
        margin-bottom: 0.8rem;
        background: linear-gradient(135deg, rgba(59,130,246,0.12), rgba(59,130,246,0.06));
        border: 1px solid rgba(59,130,246,0.25);
        border-radius: 50px;
        padding: 0.35rem 1rem;
        position: relative;
        z-index: 1;
    }}
    .stage-header-title {{
        font-size: 1.65rem;
        font-weight: 900;
        color: {C['text']};
        margin-bottom: 0.6rem;
        position: relative;
        z-index: 1;
        letter-spacing: -0.02em;
    }}
    .stage-header-desc {{
        font-size: 0.95rem;
        color: {C['text2']};
        line-height: 1.7;
        max-width: 700px;
        position: relative;
        z-index: 1;
    }}

    .step-item {{
        display: flex;
        align-items: center;
        gap: 1.2rem;
        background: linear-gradient(135deg, {C['bg2']}, {C['bg1']});
        border: 1px solid {C['border']};
        border-radius: 14px;
        padding: 1.1rem 1.4rem;
        margin-bottom: 0.7rem;
        transition: all 0.25s ease;
    }}
    .step-item:hover {{
        border-color: {C['accent']}50;
        background: linear-gradient(135deg, rgba(59,130,246,0.04), {C['bg1']});
        transform: translateX(-4px);
        box-shadow: 0 4px 12px rgba(59,130,246,0.15);
    }}
    .step-num {{
        width: 36px; height: 36px;
        border-radius: 10px;
        background: linear-gradient(135deg, {C['accent']}, {C['accent2']});
        color: white; font-weight: 800;
        font-size: 0.88rem;
        display: flex; align-items: center;
        justify-content: center; flex-shrink: 0;
        box-shadow: 0 2px 8px rgba(59,130,246,0.25);
    }}
    .step-icon {{ font-size: 1.2rem; flex-shrink: 0; }}
    .step-content {{ flex: 1; }}
    .step-title {{ font-weight: 700; color: {C['text']}; font-size: 0.95rem; }}
    .step-desc {{ font-size: 0.82rem; color: {C['text3']}; margin-top: 0.2rem; }}
    .step-badge {{
        font-size: 0.68rem;
        padding: 0.2rem 0.6rem;
        border-radius: 20px;
        font-weight: 700;
        white-space: nowrap;
        text-transform: uppercase;
        letter-spacing: 0.04em;
    }}
    .badge-new {{
        background: linear-gradient(135deg, rgba(16,185,129,0.15), rgba(16,185,129,0.08));
        color: {C['green']};
        border: 1px solid rgba(16,185,129,0.25);
    }}
    .badge-advanced {{
        background: linear-gradient(135deg, rgba(139,92,246,0.15), rgba(139,92,246,0.08));
        color: {C['accent3']};
        border: 1px solid rgba(139,92,246,0.25);
    }}
    .badge-bilingual {{
        background: linear-gradient(135deg, rgba(6,182,212,0.15), rgba(6,182,212,0.08));
        color: {C['cyan']};
        border: 1px solid rgba(6,182,212,0.25);
    }}

    .alert {{
        border-radius: 12px;
        padding: 1rem 1.3rem;
        margin: 0.8rem 0;
        font-weight: 600;
        font-size: 0.9rem;
        display: flex;
        align-items: center;
        gap: 0.8rem;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }}
    .alert-info {{
        background: linear-gradient(135deg, rgba(59,130,246,0.12), rgba(59,130,246,0.06));
        border: 1px solid rgba(59,130,246,0.25);
        color: {C['accent']};
    }}
    .alert-success {{
        background: linear-gradient(135deg, rgba(16,185,129,0.12), rgba(16,185,129,0.06));
        border: 1px solid rgba(16,185,129,0.25);
        color: {C['green']};
    }}
    .alert-warning {{
        background: linear-gradient(135deg, rgba(245,158,11,0.12), rgba(245,158,11,0.06));
        border: 1px solid rgba(245,158,11,0.25);
        color: {C['yellow']};
    }}
    .alert-danger {{
        background: linear-gradient(135deg, rgba(239,68,68,0.12), rgba(239,68,68,0.06));
        border: 1px solid rgba(239,68,68,0.25);
        color: {C['red']};
    }}

    .quality-card {{
        background: linear-gradient(135deg, {C['bg2']}, {C['bg1']});
        border: 1px solid {C['border']};
        border-radius: 18px;
        padding: 2.2rem;
        text-align: center;
        box-shadow: 0 8px 24px rgba(0,0,0,0.15);
    }}
    .quality-score {{
        font-size: 4.2rem;
        font-weight: 900;
        line-height: 1;
        background: linear-gradient(135deg, {C['accent']}, {C['cyan']});
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }}
    .quality-bar-wrap {{
        background: {C['bg3']};
        border-radius: 10px;
        height: 8px;
        overflow: hidden;
        margin: 1.2rem auto;
        max-width: 180px;
        box-shadow: inset 0 2px 4px rgba(0,0,0,0.2);
    }}
    .quality-bar-fill {{ height: 100%; border-radius: 10px; transition: all 0.4s ease; }}

    .col-profile {{
        background: linear-gradient(135deg, {C['bg2']}, {C['bg1']});
        border: 1px solid {C['border']};
        border-radius: 14px;
        padding: 1.3rem 1.5rem;
        margin-bottom: 0.8rem;
        transition: all 0.25s ease;
        box-shadow: 0 2px 6px rgba(0,0,0,0.05);
    }}
    .col-profile:hover {{ 
        border-color: {C['border2']};
        transform: translateX(-2px);
        box-shadow: 0 4px 12px rgba(59,130,246,0.1);
    }}
    .col-profile-head {{
        display: flex;
        align-items: center;
        gap: 0.6rem;
        margin-bottom: 0.8rem;
    }}
    .col-type-pill {{
        font-size: 0.68rem;
        padding: 0.18rem 0.55rem;
        border-radius: 20px;
        font-weight: 700;
        flex-shrink: 0;
    }}
    .pill-num  {{ background:rgba(6,182,212,0.15);  color:{C['cyan']};    }}
    .pill-text {{ background:rgba(99,102,241,0.15);  color:{C['accent2']}; }}
    .pill-date {{ background:rgba(16,185,129,0.15);  color:{C['green']};   }}
    .pill-bool {{ background:rgba(245,158,11,0.15);  color:{C['yellow']};  }}
    .pill-mixed{{ background:rgba(139,92,246,0.15);  color:{C['accent3']}; }}
    .col-name  {{ font-weight:800; font-size:0.97rem; color:{C['text']};   }}

    .lang-indicator {{
        display: inline-flex;
        align-items: center;
        gap: 0.3rem;
        font-size: 0.65rem;
        padding: 0.1rem 0.45rem;
        border-radius: 10px;
        font-weight: 700;
    }}
    .lang-ar {{
        background: rgba(59,130,246,0.12);
        color: {C['accent']};
        border: 1px solid rgba(59,130,246,0.2);
    }}
    .lang-en {{
        background: rgba(16,185,129,0.12);
        color: {C['green']};
        border: 1px solid rgba(16,185,129,0.2);
    }}
    .lang-mixed {{
        background: rgba(139,92,246,0.12);
        color: {C['accent3']};
        border: 1px solid rgba(139,92,246,0.2);
    }}

    .col-stats {{
        display: flex;
        gap: 1.3rem;
        flex-wrap: wrap;
        font-size: 0.82rem;
        margin-bottom: 0.7rem;
    }}
    .stat-item {{ display:flex; flex-direction:column; gap:0.1rem; }}
    .stat-label {{
        color:{C['text3']};
        font-size:0.7rem;
        font-weight:600;
        text-transform:uppercase;
    }}
    .stat-value {{ color:{C['text']}; font-weight:700; }}
    .mini-bar-wrap {{
        background: {C['bg3']};
        border-radius:6px;
        height:4px;
        overflow:hidden;
        margin-top:0.5rem;
    }}
    .mini-bar-fill {{ height:100%; border-radius:6px; }}
    .top-vals {{
        display:flex;
        flex-wrap:wrap;
        gap:0.35rem;
        margin-top:0.6rem;
    }}
    .val-chip {{
        background: {C['bg3']};
        border: 1px solid {C['border2']};
        border-radius:6px;
        padding: 0.18rem 0.55rem;
        font-size:0.74rem;
        color:{C['text2']};
        font-family: 'Tajawal', 'Inter', sans-serif;
    }}
    .val-chip strong {{ color:{C['accent']}; }}

    .miss-row {{ margin-bottom: 0.65rem; }}
    .miss-row-head {{
        display: flex;
        justify-content: space-between;
        font-size: 0.83rem;
        margin-bottom: 0.28rem;
    }}
    .miss-col-name {{ color:{C['text']}; font-weight:600; }}
    .miss-pct {{ font-weight:700; }}
    .miss-bar-wrap {{ background:{C['bg3']}; border-radius:6px; height:5px; overflow:hidden; }}
    .miss-bar-fill {{ height:100%; border-radius:6px; }}

    .impact-grid {{
        display:grid;
        grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
        gap:1rem;
        margin:1.2rem 0;
    }}
    .impact-card {{
        background: linear-gradient(135deg, {C['bg2']}, {C['bg1']});
        border:1px solid {C['border']};
        border-radius:14px;
        padding:1.3rem;
        text-align:center;
        transition: all 0.3s ease;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }}
    .impact-card:hover {{
        transform: translateY(-3px);
        box-shadow: 0 6px 16px rgba(59,130,246,0.1);
        border-color: {C['border2']};
    }}
    .impact-icon {{ font-size:1.6rem; margin-bottom:0.5rem; }}
    .impact-val {{
        font-size:1.8rem;
        font-weight:900;
        color:{C['accent']};
        line-height:1;
        margin-bottom:0.3rem;
    }}
    .impact-lbl {{
        font-size:0.75rem;
        color:{C['text3']};
        font-weight:600;
        text-transform:uppercase;
        letter-spacing:0.04em;
    }}
    .impact-sub {{ font-size:0.72rem; color:{C['text3']}; margin-top:0.2rem; }}

    .sb-logo {{
        padding: 1.4rem 1.2rem 0.9rem;
        border-bottom: 1px solid {C['border']};
        margin-bottom: 0.8rem;
    }}
    .sb-logo-title {{
        font-size:1.2rem;
        font-weight:900;
        color:{C['text']};
        display:flex;
        align-items:center;
        gap:0.5rem;
    }}
    .sb-logo-sub {{ font-size:0.76rem; color:{C['text3']}; margin-top:0.25rem; }}
    .sb-section {{ padding: 0.7rem 1.2rem; margin-bottom: 0.4rem; }}
    .sb-section-title {{
        font-size:0.7rem;
        font-weight:700;
        color:{C['text3']};
        text-transform:uppercase;
        letter-spacing:0.1em;
        margin-bottom:0.6rem;
    }}
    .sb-stat-row {{
        display:flex;
        justify-content:space-between;
        align-items:center;
        padding:0.45rem 0;
        border-bottom:1px solid {C['border']};
        font-size:0.83rem;
    }}
    .sb-stat-row:last-child {{ border-bottom:none; }}
    .sb-stat-label {{ color:{C['text2']}; }}
    .sb-stat-value {{ font-weight:700; color:{C['text']}; }}

    .log-item {{
        display:flex;
        align-items:flex-start;
        gap:0.6rem;
        padding:0.55rem 0.8rem;
        border-radius:8px;
        background:{C['bg3']};
        border:1px solid {C['border']};
        margin-bottom:0.35rem;
        font-size:0.84rem;
        color:{C['text2']};
    }}
    .log-check {{ color:{C['green']}; font-weight:900; flex-shrink:0; }}

    .feature-grid {{
        display:grid;
        grid-template-columns: repeat(auto-fit, minmax(210px, 1fr));
        gap:1.1rem;
        margin-top:2rem;
    }}
    .feature-card {{
        background: linear-gradient(135deg, {C['bg2']}, {C['bg1']});
        border:1px solid {C['border']};
        border-radius:16px;
        padding:1.5rem;
        transition:all 0.3s ease;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
    }}
    .feature-card:hover {{
        border-color:{C['accent']}50;
        transform:translateY(-4px);
        box-shadow:0 12px 30px rgba(59,130,246,0.2);
    }}
    .feature-icon {{ font-size:2rem; margin-bottom:0.8rem; display:block; }}
    .feature-title {{ font-weight:800; font-size:0.96rem; color:{C['text']}; margin-bottom:0.35rem; }}
    .feature-desc {{ font-size:0.8rem; color:{C['text3']}; line-height:1.7; }}
    .divider {{
        height:2px;
        background: linear-gradient(90deg, transparent, {C['border2']}, transparent);
        margin:1.8rem 0;
        border-radius: 1px;
    }}

    /* Responsive layout for tablets and phones */
    @media (max-width: 1280px) {{
        .block-container {{ padding: 1.6rem 1.2rem !important; }}
        section[data-testid="stSidebar"] {{ width: min(320px, 100%) !important; }}
        [data-testid="stSidebar"] {{ border-left: none !important; border-top: 1px solid {C['border']} !important; }}
        .stTabs [data-baseweb="tab"] {{ padding: 0.65rem 1rem !important; font-size: 0.95rem !important; }}
        .hero {{ padding: 2.2rem 1.2rem 1.4rem; }}
        .stage-header, .stepper-wrap, .metric-card, .impact-card, .feature-card, .quality-card, .col-profile {{ padding: 1.3rem 1rem !important; }}
    }}
    @media (max-width: 950px) {{
        .block-container {{ padding: 1.2rem 1rem !important; }}
        .hero-title {{ font-size: 2.4rem !important; }}
        .hero-subtitle {{ font-size: 1rem !important; max-width: 100% !important; }}
        .stepper-wrap {{ flex-direction: column; gap: 1rem; padding: 1rem 1rem !important; }}
        .stepper-line {{ max-width: 50px; }}
        .step-item {{ flex-direction: column; align-items: stretch; gap: 0.75rem; }}
        .col-stats {{ flex-direction: column; gap: 0.7rem; }}
        [data-testid="stDataFrame"] {{ overflow-x: auto !important; }}
        [data-testid="stFileUploader"] {{ padding: 1.8rem 1rem !important; }}
    }}
    @media (max-width: 700px) {{
        html, body, [data-testid="stAppViewContainer"], [data-testid="stMain"], .main, .block-container {{ font-size: 0.95rem !important; }}
        .block-container {{ padding: 1rem 0.8rem !important; }}
        .hero-title {{ font-size: 2rem !important; }}
        .hero-subtitle {{ font-size: 0.95rem !important; }}
        .stButton > button, .stDownloadButton > button {{ padding: 0.75rem 1rem !important; }}
        .sb-logo, .sb-section {{ padding: 0.9rem 0.95rem !important; }}
        .stage-header {{ padding: 1.2rem 0.95rem !important; }}
        .stepper-wrap {{ padding: 0.9rem 0.8rem !important; }}
        .stepper-label {{ font-size: 0.8rem !important; }}
        .step-title {{ font-size: 0.9rem !important; }}
        .metric-card, .impact-card, .feature-card {{ padding: 1rem !important; }}
        .stage-header-desc, .hero-subtitle {{ line-height: 1.6 !important; }}
    }}

    #MainMenu {{ visibility:hidden; }}
    footer {{ visibility:hidden; }}
    header {{ visibility:hidden; }}
    </style>
    """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ═══════════════════════════════════════════════════════════════════════════════
def init_state():
    defaults = {
        "df": None, "df_original": None,
        "step": 1, "cleaning_done": False,
        "cleaning_time": 0.0,
        "file_name": "", "file_size": 0,
        "rows_deleted": 0, "cells_fixed": 0,
        "dup_cols_removed": 0,
        "rows_before": 0, "rows_after": 0,
        "dates_fixed": 0, "outliers_flagged": 0,
        "types_fixed": 0,
        "export_format": "Excel",
        "profile": None,
        "quality_before": 0,
        "quality_after": 0,
        "cleaning_log": [],
        "detected_langs": {},
        # ─── Outlier Detection ───
        "selected_outlier_columns": [],
        # ─── Advanced Features ───
        "advanced_options": {
            "clean_phones": bool(AdvancedCleaner),
            "clean_numeric": bool(AdvancedCleaner),
            "fuzzy_match": bool(AdvancedCleaner),
            "fix_dates": bool(AdvancedCleaner),
            "optimize_memory": bool(AdvancedCleaner),
            "impute_text": bool(AdvancedCleaner),
            "auto_drop_low_variance": False,
        },
        "audit_trail": None,
        "advanced_log": {},
        "audit_trail_obj": None,
        "median_imputed_cells": [],
        "green_coordinates": [],
        "yellow_coordinates": [],
        "red_coordinates": [],
        "outlier_coords": [],
        "text_imputation_mask": None,  # NEW: Boolean DataFrame for imputed text cells
        "sidebar_visible": True,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


# ═══════════════════════════════════════════════════════════════════════════════
# ████████████  LANGUAGE DETECTION ENGINE  ████████████████████████████████████
# ═══════════════════════════════════════════════════════════════════════════════
ARABIC_PATTERN  = re.compile(r'[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF]+')
ENGLISH_PATTERN = re.compile(r'[A-Za-z]+')

def detect_text_language(text: str) -> str:
    """
    كشف لغة النص:
    - 'ar'    → عربي بحت
    - 'en'    → إنجليزي بحت
    - 'mixed' → خليط
    - 'other' → أرقام / رموز
    """
    if not isinstance(text, str) or not text.strip():
        return "other"
    has_ar = bool(ARABIC_PATTERN.search(text))
    has_en = bool(ENGLISH_PATTERN.search(text))
    if has_ar and has_en: return "mixed"
    if has_ar:            return "ar"
    if has_en:            return "en"
    return "other"

def detect_column_language(series: pd.Series) -> str:
    """كشف اللغة السائدة في عمود كامل"""
    non_null = series.dropna().astype(str)
    non_null = non_null[~non_null.str.lower().isin(MISSING_VALUES)]
    if len(non_null) == 0:
        return "other"
    sample = non_null.head(50)
    counts = {"ar": 0, "en": 0, "mixed": 0, "other": 0}
    for v in sample:
        counts[detect_text_language(v)] += 1
    total = sum(counts.values())
    if total == 0:
        return "other"
    ar_pct = counts["ar"] / total
    en_pct = counts["en"] / total
    if ar_pct >= 0.6:  return "ar"
    if en_pct >= 0.6:  return "en"
    if ar_pct + counts["mixed"]/total >= 0.4 and en_pct + counts["mixed"]/total >= 0.4:
        return "mixed"
    if ar_pct > en_pct: return "ar"
    if en_pct > ar_pct: return "en"
    return "mixed"


# ═══════════════════════════════════════════════════════════════════════════════
# ████████████  MISSING VALUE DETECTION  ██████████████████████████████████████
# ═══════════════════════════════════════════════════════════════════════════════
def _is_missing(val) -> bool:
    if val is None: return True
    try:
        if pd.isna(val): return True
    except Exception:
        pass
    if isinstance(val, str):
        return val.strip().lower() in MISSING_VALUES
    return False


# ═══════════════════════════════════════════════════════════════════════════════
# ████████████  DATE ENGINE  ██████████████████████████████████████████████████
# ═══════════════════════════════════════════════════════════════════════════════
EXCEL_EPOCH  = datetime(1899, 12, 30)
DATE_FORMATS = [
    "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
    "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y",
    "%d %m %Y", "%Y %m %d",
    "%d/%m/%y", "%m/%d/%y",
    "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S", "%d-%m-%Y %H:%M:%S",
    "%B %d, %Y", "%b %d, %Y",
    "%d %B %Y", "%d %b %Y",
]

def _excel_serial_to_date(serial) -> str:
    try:
        n = int(float(str(serial).strip()))
        if 1 <= n <= 2958465:
            return (EXCEL_EPOCH + pd.Timedelta(days=n)).strftime("%Y-%m-%d")
    except Exception:
        pass
    return ""

def _parse_date_string(val: str) -> str:
    val = str(val).strip()
    if not val or val.lower() in MISSING_VALUES:
        return ""
        
    # Try dateutil parser
    try:
        from dateutil import parser as du_parser
        try:
            parsed = du_parser.parse(val)
        except Exception:
            parsed = du_parser.parse(val, fuzzy=True)
        if parsed:
            return parsed.strftime("%Y-%m-%d")
    except Exception:
        pass
        
    # Fallback to pandas
    try:
        parsed = pd.to_datetime(val, dayfirst=True, errors="raise")
        if pd.notna(parsed):
            return parsed.strftime("%Y-%m-%d")
    except Exception:
        pass
        
    # Fallback to format list
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""

def _convert_cell_to_date(val) -> str:
    if val is None: return ""
    try:
        if pd.isna(val): return ""
    except Exception:
        pass
    if isinstance(val, (pd.Timestamp, datetime)):
        try: return val.strftime("%Y-%m-%d")
        except: return ""
    if isinstance(val, date) and not isinstance(val, datetime):
        try: return val.strftime("%Y-%m-%d")
        except: return ""
    if isinstance(val, (int, np.integer)):
        return _excel_serial_to_date(val)
    if isinstance(val, (float, np.floating)):
        if np.isnan(val): return ""
        if 1 <= val <= 2958465:
            return _excel_serial_to_date(int(val))
        return ""
    if isinstance(val, str):
        s = val.strip()
        if not s or s.lower() in MISSING_VALUES: return ""
        if re.fullmatch(r"\d{4,6}", s):
            r = _excel_serial_to_date(s)
            if r: return r
        return _parse_date_string(s)
    return ""

def _is_date_column(series: pd.Series) -> bool:
    if pd.api.types.is_datetime64_any_dtype(series): return True
    date_kws = [
        "date", "تاريخ", "يوم", "day", "time", "وقت",
        "created", "updated", "birth", "ميلاد",
        "انشاء", "اضافة", "تعديل", "انتهاء",
        "بداية", "نهاية", "timestamp", "recorded",
        "created_at", "updated_at", "birth_date",
    ]
    col_lower = str(series.name).lower()
    identifier_kws = [
        "id", "order id", "orderid", "order_no", "order no", "order number",
        "رقم", "رقم الطلب", "رقم الطلبية", "كود", "sku", "barcode", "token", "code",
        "uid", "user id",
    ]
    if any(k in col_lower for k in date_kws): return True
    if any(k in col_lower for k in identifier_kws): return False
    if series.dtype == object:
        non_null = series.dropna().astype(str)
        non_null = non_null[~non_null.str.lower().isin(MISSING_VALUES)]
        if len(non_null) == 0: return False
        avg_len = non_null.str.len().mean()
        if 4 <= avg_len <= 25:
            sample = non_null.head(30)
            hits   = sum(1 for v in sample if _parse_date_string(v) != "")
            if hits / len(sample) >= 0.6: return True
    return False

def fix_date_columns(df: pd.DataFrame):
    total = 0
    for col in df.columns:
        if not _is_date_column(df[col]): continue
        new_vals, cnt = [], 0
        for val in df[col]:
            if _is_missing(val):
                new_vals.append(val); continue
            converted = _convert_cell_to_date(val)
            if converted:
                new_vals.append(converted); cnt += 1
            else:
                new_vals.append(val)
        df[col] = new_vals
        total  += cnt
    return df, total


# ═══════════════════════════════════════════════════════════════════════════════
# ████████████  TEXT NORMALIZATION — ARABIC + ENGLISH  ████████████████████████
# ═══════════════════════════════════════════════════════════════════════════════
def _normalize_arabic(val: str) -> str:
    """تطبيع النص العربي فقط"""
    # إزالة التشكيل باستخدام تعبير نمطي سريع
    arabic_diacritics = re.compile(r'[\u064B-\u0652\u0670]')
    val = arabic_diacritics.sub('', val)
    # أحرف خاصة
    for ch in "\u0640\u200C\u200D\u200E\u200F\uFEFF":
        val = val.replace(ch, "")
    # توحيد الألف
    for ch in "أإآٱ":
        val = val.replace(ch, "ا")
    # التاء المربوطة والياء
    val = val.replace("ة", "ه").replace("ى", "ي")
    return val

def _normalize_english(val: str) -> str:
    """تطبيع النص الإنجليزي"""
    # إزالة الرموز غير الضرورية من الطرفين
    val = val.strip(".,;:!?\"'()[]{}")
    # توحيد المسافات
    val = re.sub(r"\s+", " ", val).strip()
    return val.title()

def _normalize_text(val, is_pii: bool = False) -> str:
    """
    تطبيع ذكي يكشف اللغة ويطبّق التطبيع المناسب:
    - عربي  → تطبيع عربي كامل
    - إنجليزي → تطبيع إنجليزي
    - مختلط → تطبيع للجزأين
    """
    if not isinstance(val, str):
        return str(val)

    # إزالة الرموز الغريبة للملفات العادية
    if not is_pii:
        for char in "@#_?":
            val = val.replace(char, " ")

    lang = detect_text_language(val)

    if lang == "ar":
        val = _normalize_arabic(val)
    elif lang == "en":
        val = _normalize_english(val)
    elif lang == "mixed":
        # نفصل العربي والإنجليزي ونطبّق على كل جزء
        parts = re.split(r'(\s+)', val)
        result = []
        for part in parts:
            if re.search(r'[\u0600-\u06FF]', part):
                result.append(_normalize_arabic(part))
            elif re.search(r'[A-Za-z]', part):
                result.append(_normalize_english(part))
            else:
                result.append(part)
        val = "".join(result)

    return re.sub(r"\s+", " ", val).strip()

def _detect_pii(series):
    # Sample up to 100 non-null values
    sample = series.dropna().head(100).astype(str).str.strip()
    if len(sample) == 0:
        return None
    
    email_rx = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    phone_rx = re.compile(r'^\+?[0-9][0-9\s\-()]{6,18}[0-9]$')
    
    emails = sample.apply(lambda x: bool(email_rx.match(x)))
    phones = sample.apply(lambda x: bool(phone_rx.match(x)) and len(re.sub(r'\D', '', x)) >= 7)
    
    if emails.mean() > 0.5:
        return "Email"
    if phones.mean() > 0.5:
        return "Phone"
    return None

def _arabic_english_normalization_pass(df: pd.DataFrame) -> None:
    """تطبيع كامل على جميع الأعمدة النصية مع استثناء وحفظ PII"""
    for col in df.select_dtypes(include=["object", "string"]).columns:
        if col.startswith("__") or _is_date_column(df[col]):
            continue
            
        pii = _detect_pii(df[col])
        if pii == "Email":
            df[col] = df[col].apply(
                lambda v: str(v).strip().lower() if pd.notna(v) else v
            )
        elif pii == "Phone":
            continue
        else:
            df[col] = df[col].apply(
                lambda v: _normalize_text(v, is_pii=False) if pd.notna(v) else v
            )



def _standardize_majority_words(df: pd.DataFrame) -> int:
    """Standardize spelling variations to the dominant word in each text column."""
    fixed = 0
    for col in df.select_dtypes(include=["object", "string"]).columns:
        if col.startswith("__"): continue
        values = df[col].dropna().astype(str).str.strip()
        values = values[~values.str.lower().isin(MISSING_VALUES)]
        if values.empty:
            continue
        mode_vals = values.mode()
        if mode_vals.empty:
            continue
        majority_word = str(mode_vals.iloc[0]).strip()
        if not majority_word:
            continue

        for idx, raw_val in df[col].items():
            if pd.isna(raw_val) or not isinstance(raw_val, str):
                continue
            cell = raw_val.strip()
            if not cell or cell == majority_word or cell.lower() in MISSING_VALUES:
                continue
            ratio = difflib.SequenceMatcher(None, cell, majority_word).ratio()
            if ratio >= 0.85:
                df.at[idx, col] = majority_word
                fixed += 1
    return fixed


# ═══════════════════════════════════════════════════════════════════════════════
# STYLER - OPTIMIZED FOR LARGE DATASETS
# ═══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
def _detect_type(s: pd.Series) -> str:
    if pd.api.types.is_bool_dtype(s):            return "boolean"
    if pd.api.types.is_datetime64_any_dtype(s):  return "date"
    if pd.api.types.is_numeric_dtype(s):         return "numeric"
    if _is_date_column(s):                       return "date"
    nn = s.dropna().astype(str)
    if len(nn) > 0:
        hits = sum(1 for v in nn.head(30)
                   if re.fullmatch(r"[-+]?\d*\.?\d+", v.strip()))
        if hits / max(len(nn.head(30)), 1) >= 0.8: return "numeric"
        uniq = set(nn.str.lower().unique())
        if uniq.issubset({"true","false","yes","no","1","0","نعم","لا"}):
            return "boolean"
    return "text"

def _quality_score(df: pd.DataFrame) -> float:
    clean = df.drop(columns=["__outlier__"], errors="ignore")
    if clean.size == 0: return 0.0
    miss = clean.isnull().sum().sum() / clean.size
    dup  = clean.duplicated().sum() / max(len(clean), 1)
    seen, dc = {}, 0
    for col in clean.columns:
        h = hashlib.md5(clean[col].astype(str).str.cat().encode()).hexdigest()
        dc += 1 if h in seen else 0
        seen[h] = col
    dcs = max(0.0, 10*(1 - dc/max(len(clean.columns),1)))
    ti  = sum(
        1 for col in clean.select_dtypes(include="object").columns
        if len(clean[col].dropna()) > 0 and
        sum(1 for v in clean[col].dropna().astype(str).head(30)
            if re.fullmatch(r"[-+]?\d*\.?\d+", v.strip()))
        / max(len(clean[col].dropna().head(30)), 1) >= 0.8
    )
    return round(
        max(0.0, 40*(1-miss)) + max(0.0, 20*(1-dup)) +
        dcs + max(0.0, 30*(1-ti/max(len(clean.columns),1))), 1
    )

def build_profile(df: pd.DataFrame) -> dict:
    """Build data profile - expensive computation, should be cached at call site"""
    report = {}
    clean  = df.drop(columns=["__outlier__"], errors="ignore")
    for col in clean.columns:
        s   = clean[col]
        nn  = s.dropna()
        tot = len(s)
        nc  = int(s.isnull().sum())
        ct  = _detect_type(s)
        # كشف لغة العمود
        col_lang = detect_column_language(s) if ct == "text" else "other"
        
        # PII Detection
        pii_type = _detect_pii(s)
        
        # Drop Recommendations
        null_pct = nc / max(tot, 1) * 100
        is_high_missing = null_pct > 90.0
        
        if len(nn) > 0:
            val_counts = nn.value_counts()
            most_freq_pct = (val_counts.iloc[0] / len(nn) * 100) if len(val_counts) > 0 else 0.0
        else:
            most_freq_pct = 0.0
            
        is_low_variance = most_freq_pct > 98.0 and len(nn) > 1
        
        entry = {
            "type": ct, "total": tot,
            "null_count": nc,
            "null_pct": round(null_pct, 1),
            "unique": int(s.nunique()),
            "fill_rate": round((tot-nc)/max(tot,1)*100, 1),
            "language": col_lang,
            "pii_type": pii_type,
            "is_high_missing": is_high_missing,
            "is_low_variance": is_low_variance,
            "recommend_drop": is_high_missing or is_low_variance
        }
        if ct == "numeric" and len(nn) > 0:
            try:
                ns = pd.to_numeric(nn, errors="coerce").dropna()
                if len(ns) > 0:
                    q1, q3 = ns.quantile(0.25), ns.quantile(0.75)
                    iqr = q3 - q1
                    entry.update({
                        "mean":     round(float(ns.mean()), 2),
                        "median":   round(float(ns.median()), 2),
                        "std":      round(float(ns.std()), 2),
                        "min":      round(float(ns.min()), 2),
                        "max":      round(float(ns.max()), 2),
                        "outliers": int(((ns<q1-1.5*iqr)|(ns>q3+1.5*iqr)).sum()),
                    })
            except Exception:
                pass
        if ct == "text" and len(nn) > 0:
            top = nn.astype(str).value_counts().head(5).to_dict()
            entry["top_values"] = {str(k):int(v) for k,v in top.items()}
        report[col] = entry
    return report


# ═══════════════════════════════════════════════════════════════════════════════
# ████████████  CLEANING PIPELINE — 10 STEPS  █████████████████████████████████
# ═══════════════════════════════════════════════════════════════════════════════
def _remove_dup_cols(df):
    seen, drop = {}, []
    for col in df.columns:
        h = hashlib.md5(df[col].astype(str).str.cat().encode()).hexdigest()
        if h in seen: drop.append(col)
        else: seen[h] = col
    if drop: df.drop(columns=drop, inplace=True)
    return df, len(drop)

def _strip_ws(df):
    for col in df.select_dtypes(include=["object","string"]).columns:
        df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)

@st.cache_data
def _flag_outliers(df, columns=None):
    """
    رصد الشاذات باستخدام خوارزمية IQR | Detect outliers using IQR algorithm
    
    Args:
        df: DataFrame
        columns: List of column names to check for outliers. 
                If None or empty, no outliers are flagged.
                If provided, only checks these columns.
    
    Returns:
        df with __outlier__ flag, count of outliers found
    """
    df   = df.copy()
    mask = pd.Series([False]*len(df), index=df.index)
    cnt  = 0
    
    # If no columns provided, default to all numeric columns (fully generic)
    if columns is None or len(columns) == 0:
        columns = df.select_dtypes(include=["number"]).columns.tolist()

    coords = []

    # Only check the provided columns
    for col in columns:
        if col.startswith("__"): continue
        if col not in df.columns: continue
        try:
            num = pd.to_numeric(df[col], errors="coerce")
            if num.notna().sum() < 4: continue
            q1, q3 = num.quantile(0.25), num.quantile(0.75)
            iqr = q3 - q1
            if iqr == 0: continue
            m   = ((num < q1-1.5*iqr) | (num > q3+1.5*iqr)) & num.notna()
            if m.any():
                cnt += int(m.sum())
                coords.extend((idx, col) for idx in df.index[m])
                mask = mask | m
        except Exception: continue
    df["__outlier__"] = mask
    return df, cnt, coords

def _std_types(df):
    cnt  = 0
    bmap = {
        "true":"True","false":"False",
        "yes":"True","no":"False",
        "1":"True","0":"False",
        "نعم":"True","لا":"False",
    }
    for col in df.columns:
        if col.startswith("__") or _is_date_column(df[col]): continue
        if df[col].dtype == object:
            try:
                df[col] = pd.to_numeric(
                    df[col].replace("غير محدد", np.nan), errors="raise"
                )
                cnt += 1; continue
            except Exception: pass
            low = df[col].dropna().astype(str).str.lower()
            if set(low.unique()).issubset(set(bmap)):
                df[col] = df[col].map(lambda x: bmap.get(str(x).lower(), x))
                cnt += 1
    return df, cnt

def _infer_imputation_guides(df: pd.DataFrame):
    columns = list(df.columns)
    fine_priority = [
        "item name", "product name", "item", "product", "sku", "variant", "model",
        "item_id", "product_id", "title",
    ]
    coarse_priority = [
        "category", "department", "segment", "group", "type", "class", "family",
        "category name", "department name",
    ]

    def choose(priorities):
        for pat in priorities:
            for col in columns:
                if col.lower() == pat or pat in col.lower():
                    return col
        return None

    return choose(fine_priority), choose(coarse_priority)


def _fill_missing(df):
    """
    ملء القيم الناقصة بذكاء:
    - معلومات رقمية ونصية → محرك ملء متقدم يعتمد على السياق
    - يحتفظ بكل القيم الأصلية ويضيف أعمدة حالة التنبؤ لكل عمود
    """
    median_imputed = []
    text_missing = []

    if AdvancedCleaner and universal_missing_value_imputation:
        numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
        categorical_cols = df.select_dtypes(include=["object", "string", "category"]).columns.tolist()
        fine, coarse = _infer_imputation_guides(df)

        if fine and coarse and (numeric_cols or categorical_cols):
            df_orig = df.copy()
            try:
                df = universal_missing_value_imputation(
                    df.copy(),
                    numeric_cols_to_fill=numeric_cols,
                    categorical_cols_to_fill=categorical_cols,
                    fine_grain_guide=fine,
                    coarse_grain_guide=coarse,
                )

                for col in numeric_cols:
                    if col in df.columns:
                        missing_mask = df_orig[col].isna() & df[col].notna()
                        median_imputed.extend((idx, col) for idx in df_orig.index[missing_mask])

                for col in categorical_cols:
                    if col in df.columns:
                        missing_mask = df_orig[col].isna() & df[col].notna()
                        text_missing.extend((idx, col) for idx in df_orig.index[missing_mask])

                return df, median_imputed, text_missing
            except Exception as e:
                logger.warning("Advanced universal_missing_value_imputation failed: %s", e)

    for col in df.columns:
        if col.startswith("__"): continue

        if pd.api.types.is_numeric_dtype(df[col]):
            missing_mask = df[col].isna()
            if missing_mask.any():
                non_null = df[col].dropna()
                if len(non_null) > 0:
                    median_val = non_null.median()
                    if not pd.isna(median_val):
                        df.loc[missing_mask, col] = median_val
                        median_imputed.extend((idx, col) for idx in df.index[missing_mask])
            continue

        if pd.api.types.is_string_dtype(df[col]) or df[col].dtype == object:
            df[col] = _drop_blank_strings_for_text(df, col)
            missing_mask = df[col].isna()
            if missing_mask.any():
                df.loc[missing_mask, col] = "غير محدد"
                text_missing.extend((idx, col) for idx in df.index[missing_mask])
            continue

    return df, median_imputed, text_missing


def context_aware_hierarchical_impute_numeric(
    df: pd.DataFrame,
    target_numeric_col: str,
    fine_grain_col: str,
    coarse_grain_col: str,
) -> pd.DataFrame:
    """Impute missing numeric values with context-aware hierarchical logic.

    This function fills missing values in `target_numeric_col` using a strict waterfall:
      1. Item median grouped by `fine_grain_col`.
      2. Category median grouped by `coarse_grain_col`.
      3. Global median across `target_numeric_col`.

    The returned DataFrame includes a tracking column named
    `f"{target_numeric_col}_imputation_status"`.
    """
    if target_numeric_col not in df.columns:
        raise KeyError(f"Missing target column: {target_numeric_col}")
    if fine_grain_col not in df.columns:
        raise KeyError(f"Missing fine-grain column: {fine_grain_col}")
    if coarse_grain_col not in df.columns:
        raise KeyError(f"Missing coarse-grain column: {coarse_grain_col}")

    df = df.copy()
    status_col = f"{target_numeric_col}_imputation_status"

    numeric_target = pd.to_numeric(df[target_numeric_col], errors="coerce")
    df[status_col] = "Original Value"

    missing_mask = numeric_target.isna()
    if not missing_mask.any():
        df[target_numeric_col] = numeric_target
        return df

    item_median = numeric_target.groupby(df[fine_grain_col]).transform("median")
    category_median = numeric_target.groupby(df[coarse_grain_col]).transform("median")
    global_median = numeric_target.median()

    fill_item = missing_mask & item_median.notna()
    if fill_item.any():
        numeric_target.loc[fill_item] = item_median.loc[fill_item]
        df.loc[fill_item, status_col] = "Filled by Item Median"

    still_missing = missing_mask & numeric_target.isna()
    fill_category = still_missing & category_median.notna()
    if fill_category.any():
        numeric_target.loc[fill_category] = category_median.loc[fill_category]
        df.loc[fill_category, status_col] = "Filled by Category Median"

    still_missing = missing_mask & numeric_target.isna()
    if still_missing.any() and pd.notna(global_median):
        numeric_target.loc[still_missing] = global_median
        df.loc[still_missing, status_col] = "Filled by Global Median"

    df[target_numeric_col] = numeric_target
    return df


def streamlit_highlight_item_median_rows(df: pd.DataFrame, target_numeric_col: str) -> object:
    """Return a Styler that highlights rows filled by item median."""
    status_col = f"{target_numeric_col}_imputation_status"
    if status_col not in df.columns:
        raise KeyError(f"Tracking column not found: {status_col}")

    def _highlight(row):
        if row[status_col] == "Filled by Item Median":
            return ["background-color: rgba(46, 204, 113, 0.2);" for _ in row]
        return ["" for _ in row]

    return df.style.apply(_highlight, axis=1)


def run_cleaning():
    import time
    start_time = time.time()
    df  = st.session_state.df.copy()
    df  = enforce_strict_numeric_candidates(df)
    log = []
    rb  = len(df)
    qs  = _quality_score(df)
    audit_trail = AuditTrail() if AdvancedCleaner else None
    st.session_state["audit_trail_obj"] = audit_trail  # Persist for subsequent operations and display
    adv_opts = st.session_state.get("advanced_options", {})

    # 1. حذف أعمدة مكررة
    df, dc = _remove_dup_cols(df)
    log.append(f"حذف {dc} عمود مكرر | Removed {dc} duplicate columns")

    # ─── 1.1 الكشف الذكي عن الأعمدة ذات التباين المنخفض أو المفقودة ───
    cols_to_drop = []
    dropped_reasons = []
    for col in df.columns:
        if col.startswith("__"): continue
        s = df[col]
        nn = s.dropna()
        tot = len(s)
        if tot == 0: continue
        nc = int(s.isnull().sum())
        null_pct = nc / tot
        
        # Check high missing (> 90%)
        if null_pct > 0.90:
            cols_to_drop.append(col)
            dropped_reasons.append(f"عمود شبه فارغ {col} ({null_pct*100:.1f}% قيم ناقصة) | empty column {col} ({null_pct*100:.1f}% missing)")
            continue
            
        # Check low variance (> 98% identical values in non-nulls)
        if len(nn) > 0:
            val_counts = nn.value_counts()
            if len(val_counts) > 0:
                most_freq_pct = val_counts.iloc[0] / len(nn)
                if most_freq_pct > 0.98 and len(nn) > 1:
                    cols_to_drop.append(col)
                    dropped_reasons.append(f"عمود ذو تباين منخفض {col} ({most_freq_pct*100:.1f}% قيمة مكررة) | low-variance column {col} ({most_freq_pct*100:.1f}% identical value)")

    if adv_opts.get("auto_drop_low_variance", False) and cols_to_drop:
        df.drop(columns=cols_to_drop, inplace=True)
        for reason in dropped_reasons:
            log.append(f"حذف تلقائي: {reason}")
    elif cols_to_drop:
        log.append("⚠️ تم العثور على أعمدة محتملة للحذف بسبب القيم الناقصة أو التباين المنخفض، لكنها لم تُحذف لأن خيار الحذف التلقائي غير مفعل.")

    # 2. تنظيف المسافات
    _strip_ws(df)
    log.append("تنظيف المسافات | Whitespace cleaned")

    # ─── ADVANCED: توحيد أرقام التليفونات ───
    if adv_opts.get("clean_phones", False) and AdvancedCleaner:
        try:
            df, phone_log = clean_phone_column(df)
            if phone_log.get("phones_standardized", 0) > 0:
                log.append(f"✨ توحيد التليفونات: {phone_log['phones_standardized']} | Phone standardization: {phone_log['phones_standardized']}")
                if audit_trail:
                    audit_trail.log_bulk_change(phone_log['phones_standardized'], "Phone number standardization")
        except Exception as e:
            log.append(f"⚠️ خطأ في توحيد التليفونات: {str(e)}")

    # ─── ADVANCED: تنظيف الأرقام المخلوطة ───
    if adv_opts.get("clean_numeric", False) and AdvancedCleaner:
        try:
            before_numeric = df.copy()
            df, num_log = clean_numeric_columns(df)
            if num_log.get("columns_cleaned", 0) > 0:
                log.append(f"✨ تنظيف الأرقام: {num_log['values_converted']} | Numeric cleaning: {num_log['values_converted']}")
                if audit_trail:
                    log_dataframe_changes(
                        audit_trail,
                        before_numeric,
                        df,
                        cols=before_numeric.columns.tolist(),
                        reason="Clean Mixed Numbers / تنظيف الأرقام المخلوطة"
                    )
        except Exception as e:
            log.append(f"⚠️ خطأ في تنظيف الأرقام: {str(e)}")

    # ─── ADVANCED: الدمج الذكي للمتشابهات ───
    if adv_opts.get("fuzzy_match", False) and AdvancedCleaner:
        try:
            before_fuzzy = df.copy()
            df, fuzzy_log = apply_fuzzy_matching(df, threshold=90)
            if fuzzy_log.get("total_changes", 0) > 0:
                log.append(f"✨ الدمج الذكي: {fuzzy_log['total_changes']} | Fuzzy matching: {fuzzy_log['total_changes']}")
                if audit_trail:
                    log_dataframe_changes(
                        audit_trail,
                        before_fuzzy,
                        df,
                        cols=before_fuzzy.columns.tolist(),
                        reason="Fuzzy matched text / تم دمج القيمة المتشابهة"
                    )
        except Exception as e:
            log.append(f"⚠️ خطأ في الدمج الذكي: {str(e)}")

    # 3. إصلاح التواريخ (العادي أو المتقدم)
    if adv_opts.get("fix_dates", False) and AdvancedCleaner:
        try:
            before_dates = df.copy()
            df, dates_log = clean_dates_advanced(df)
            dates = dates_log.get("dates_fixed", 0)
            log.append(f"✨ تصحيح التواريخ المتقدم: {dates} | Advanced date fixing: {dates}")
            if audit_trail:
                log_dataframe_changes(
                    audit_trail,
                    before_dates,
                    df,
                    cols=before_dates.columns.tolist(),
                    reason="Advanced Date Validation / التحقق المتقدم من التواريخ"
                )
        except Exception as e:
            log.append(f"⚠️ خطأ في تصحيح التواريخ: {str(e)}")
            df, dates = fix_date_columns(df)
            log.append(f"تصحيح التواريخ: {dates} | Date fixed: {dates}")
    else:
        df, dates = fix_date_columns(df)
        log.append(f"تصحيح التواريخ → YYYY-MM-DD | Fixed {dates} dates")

    # 4. تطبيع النصوص (عربي + إنجليزي)
    _arabic_english_normalization_pass(df)
    log.append("تطبيع النصوص العربية والإنجليزية | AR+EN text normalized")

    # 4.1 توحيد الكلمات السائدة في الأعمدة النصية باستخدام أغلبية القيم
    std_cells = _standardize_majority_words(df)
    cells = std_cells
    if std_cells > 0:
        log.append(f"✨ توحيد الكلمات السائدة: {std_cells} خلية | Majority word standardization: {std_cells} cells")

    # Note: removed header/keyword-based inference to keep pipeline schema-agnostic

    smart_cells = 0

    # 6. توحيد الأنواع
    df, types = _std_types(df)
    log.append(f"توحيد الأنواع: {types} عمود | Type standardization: {types} cols")

    # 7. رصد الشاذات (آلي - يتم تطبيقه على كل الأعمدة الرقمية)
    numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
    text_cols = df.select_dtypes(include=["object", "string", "category"]).columns.tolist()

    st.session_state["green_coordinates"] = _collect_missing_numeric_coords(df, numeric_cols)
    st.session_state["red_coordinates"] = _collect_missing_text_coords(df, text_cols)

    selected_cols = st.session_state.get("selected_outlier_columns", [])
    cols_for_outliers = selected_cols if selected_cols else None
    df, outl, outlier_coords = _flag_outliers(df, columns=cols_for_outliers)
    st.session_state["yellow_coordinates"] = list(outlier_coords)
    st.session_state["outlier_coords"] = list(outlier_coords)
    if outl > 0:
        cols_flagged = len(set(c for _, c in outlier_coords))
        log.append(f"رصد {outl} قيمة شاذة في {cols_flagged} أعمدة رقمية | {outl} outliers flagged across {cols_flagged} numeric columns")
    else:
        log.append("لا توجد قيم شاذة رقمية مرصودة | No numeric outliers detected")

    # 8. حذف مكررات
    df.drop_duplicates(
        subset=[c for c in df.columns if not c.startswith("__")],
        inplace=True, keep="first"
    )
    ra = len(df)
    log.append(f"حذف {rb-ra} صف مكرر | Removed {rb-ra} duplicate rows")

    # 9. ملء الفراغات بذكاء
    df, median_imputed, text_missing = _fill_missing(df)
    # Drop internal imputation tracking columns before output/export
    status_cols = [c for c in df.columns if str(c).endswith("_imputation_status")]
    if status_cols:
        df = df.drop(columns=status_cols, errors="ignore")

    st.session_state["median_imputed_cells"] = list(median_imputed)
    st.session_state["green_coordinates"] = list(median_imputed)
    st.session_state["red_coordinates"] = list(text_missing)
    log.append("ملء القيم الفارغة بذكاء | Smart fill")

    # ─── NEW: DYNAMIC CONTEXTUAL TEXT IMPUTATION ───
    text_imputation_mask = pd.DataFrame(False, index=df.index, columns=df.columns)
    if adv_opts.get("impute_text", False) and AdvancedCleaner:
        try:
            from core.text_processor import dynamic_contextual_text_imputation
            before_text_impute = df.copy()
            df, text_imputation_mask = dynamic_contextual_text_imputation(df)
            imputation_count = int(text_imputation_mask.sum().sum())
            if imputation_count > 0:
                log.append(f"✨ استنتاج نصوص سياقي: {imputation_count} خلية | Dynamic contextual text imputation: {imputation_count} cells")
                if audit_trail:
                    log_dataframe_changes(
                        audit_trail,
                        before_text_impute,
                        df,
                        cols=before_text_impute.columns.tolist(),
                        reason="Dynamic Contextual Text Imputation / استنتاج النصوص السياقي"
                    )
            st.session_state["text_imputation_mask"] = text_imputation_mask
        except Exception as e:
            log.append(f"⚠️ خطأ في استنتاج النصوص السياقي: {str(e)[:80]}")
            st.session_state["text_imputation_mask"] = None

    # ─── ADVANCED: تحسين الذاكرة ───
    mem_log = {}
    if adv_opts.get("optimize_memory", False) and AdvancedCleaner:
        try:
            df, mem_log = optimize_memory(df)
            log.append(f"✨ تحسين الذاكرة: {mem_log['reduction_percent']}% | Memory optimization: {mem_log['reduction_percent']}%")
        except Exception as e:
            log.append(f"⚠️ خطأ في تحسين الذاكرة: {str(e)}")

    # 10. إعادة حساب الجودة
    qe = _quality_score(df)
    log.append(f"الجودة | Quality: {qs} → {qe} / 100")

    elapsed_time = time.time() - start_time

    st.session_state.update({
        "df": df, "cleaning_done": True, "step": 4,
        "cleaning_time": elapsed_time,
        "rows_deleted": rb-ra, "cells_fixed": cells + smart_cells,
        "dates_fixed": dates, "dup_cols_removed": dc,
        "rows_before": rb, "rows_after": ra,
        "outliers_flagged": outl, "types_fixed": types,
        "profile": build_profile(df),
        "quality_before": qs, "quality_after": qe,
        "cleaning_log": log,
        "audit_trail": audit_trail.to_dataframe() if audit_trail else None,
        "advanced_log": {
            "memory": mem_log,
        },
    })


# ═══════════════════════════════════════════════════════════════════════════════
# ████████████  FILE READER  ██████████████████████████████████████████████████
# ═══════════════════════════════════════════════════════════════════════════════
@st.cache_data
def read_file_cached(file_name: str, file_bytes: bytes) -> pd.DataFrame:
    kw   = dict(dtype=str, keep_default_na=False)
    name = file_name.lower()

    if name.endswith(".csv"):
        # جرب encodings متعددة
        encodings = ["utf-8", "utf-8-sig", "windows-1256",
                     "windows-1252", "latin-1", "cp1256"]
        df = None
        for enc in encodings:
            try:
                df = pd.read_csv(io.BytesIO(file_bytes), encoding=enc, **kw)
                break
            except (UnicodeDecodeError, Exception):
                continue
        if df is None:
            df = pd.read_csv(io.BytesIO(file_bytes), encoding="latin-1", **kw)

    elif name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl", **kw)
    else:
        raise ValueError(f"نوع الملف غير مدعوم | Unsupported file type: {name}")

    # تنظيف أسماء الأعمدة
    df.columns = [str(c).strip() for c in df.columns]

    # استبدال NA
    na_vals = [
        "", "nan", "NaN", "NaT", "None", "none",
        "null", "NULL", "N/A", "n/a", "#N/A", "NA",
    ]
    df.replace(na_vals, np.nan, inplace=True)
    return df

def read_file(uploaded_file) -> pd.DataFrame:
    return read_file_cached(uploaded_file.name, uploaded_file.getvalue())


NUMERIC_FIELD_KEYWORDS = [
    "price", "amount", "total", "cost", "fee", "tax", "value",
    "balance", "sum", "subtotal", "net", "gross", "paid", "charge",
    "invoice", "discount", "salary", "commission", "budget",
]
ARABIC_NUMERIC_FIELD_KEYWORDS = [
    "سعر", "مبلغ", "تكلفة", "قيمة", "إجمالي", "المجموع", "رصيد",
    "دفع", "فاتورة", "ضريبة", "رسوم", "صاف", "دفعة",
]

def _is_financial_column_name(col_name: str) -> bool:
    lower = str(col_name).lower()
    return any(keyword in lower for keyword in NUMERIC_FIELD_KEYWORDS) or any(keyword in lower for keyword in ARABIC_NUMERIC_FIELD_KEYWORDS)


def enforce_strict_numeric_candidates(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        if _is_financial_column_name(col):
            coerced = pd.to_numeric(df[col], errors="coerce")
            if coerced.notna().eq(df[col].notna()).all():
                df[col] = coerced
    return df


def _collect_missing_numeric_coords(df: pd.DataFrame, cols) -> list:
    return [(idx, col) for col in cols for idx in df.index[df[col].isna()]]


def _collect_missing_text_coords(df: pd.DataFrame, cols) -> list:
    coords = []
    for col in cols:
        mask = df[col].isna() | df[col].astype(str).str.strip().eq("")
        for idx in df.index[mask]:
            coords.append((idx, col))
    return coords


def _drop_blank_strings_for_text(df: pd.DataFrame, col: str) -> pd.Series:
    series = df[col]
    if pd.api.types.is_string_dtype(series) or series.dtype == object:
        return series.where(~series.astype(str).str.strip().eq(""), np.nan)
    return series


# ═══════════════════════════════════════════════════════════════════════════════
# ████████████  EXPORT ENGINE  ████████████████████████████████████████████████
# ═══════════════════════════════════════════════════════════════════════════════
def _safe_export_convert(df: pd.DataFrame) -> pd.DataFrame:
    """Safely convert nullable ints / categories for export."""
    de = df.copy()
    for col in de.columns:
        if pd.api.types.is_integer_dtype(de[col]) and str(de[col].dtype).startswith("Int"):
            # Nullable integer (e.g., Int32) → standard float to preserve NaN
            de[col] = de[col].astype("float64")
        elif pd.api.types.is_categorical_dtype(de[col]):
            de[col] = de[col].astype(str)
    return de


def export_excel(df: pd.DataFrame) -> bytes:
    de = _safe_export_convert(df)
    oflg = de.pop("__outlier__") if "__outlier__" in de.columns else None

    date_cols = {
        col for col in df.columns
        if not col.startswith("__") and _is_date_column(df[col])
    }

    red_coords = set(tuple(c) for c in st.session_state.get("red_coordinates", []))
    green_coords = set(tuple(c) for c in st.session_state.get("green_coordinates", []))
    yellow_coords = set(tuple(c) for c in st.session_state.get("yellow_coordinates", []))
    numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
    col_langs = {}
    for col in df.columns:
        if col.startswith("__"): continue
        col_langs[col] = detect_column_language(df[col])

    out = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Data | البيانات المنظفة"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="1D4ED8")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    missing_fill = PatternFill(fill_type="solid", fgColor="F8D7DA")
    outlier_fill = PatternFill(fill_type="solid", fgColor="FFF3CD")
    median_fill = PatternFill(fill_type="solid", fgColor="D4EDDA")

    for ci, col in enumerate(de.columns, start=1):
        header = ws.cell(row=1, column=ci, value=col)
        header.font = header_font
        header.fill = header_fill
        header.alignment = align_center
        header.border = border
        width = min(max(
            de[col].astype(str).fillna("").apply(lambda x: len(str(x))).max() if len(de) > 0 else 10,
            len(str(col))
        ) + 4, 45)
        ws.column_dimensions[get_column_letter(ci)].width = width

    for ri, idx in enumerate(de.index, start=2):
        actual_idx = df.index[ri - 2]
        for ci, col in enumerate(de.columns, start=1):
            raw_val = de.at[idx, col]
            coord = (actual_idx, col)
            is_missing_text = coord in red_coords
            is_median = coord in green_coords
            is_outlier = coord in yellow_coords
            lang = col_langs.get(col, "ar")

            cell = ws.cell(row=ri, column=ci)
            cell.border = border
            if col in date_cols:
                cell.alignment = align_center
            elif lang == "en":
                cell.alignment = align_left
            elif lang == "mixed":
                cell.alignment = align_center
            else:
                cell.alignment = align_right

            if is_median:
                cell.fill = median_fill
            elif is_outlier:
                cell.fill = outlier_fill
            elif is_missing_text:
                cell.fill = missing_fill

            if is_missing_text:
                cell.value = "غير محدد"
                cell.number_format = "@"
                continue

            if col in date_cols:
                cell.value = "" if pd.isna(raw_val) else str(raw_val)
                cell.number_format = "@"
                continue

            if col in numeric_cols:
                if pd.isna(raw_val) or str(raw_val).strip() == "":
                    cell.value = None
                elif isinstance(raw_val, (int, np.integer)):
                    cell.value = int(raw_val)
                    cell.number_format = "#,##0"
                elif isinstance(raw_val, (float, np.floating)):
                    cell.value = float(raw_val)
                    cell.number_format = "#,##0.##"
                else:
                    try:
                        numeric_val = float(str(raw_val))
                        cell.value = numeric_val
                        cell.number_format = "#,##0.##"
                    except Exception:
                        cell.value = str(raw_val)
                        cell.number_format = "@"
                continue

            if pd.isna(raw_val):
                raw_val = None

            cell.value = raw_val
            cell.number_format = "@"

    max_col = get_column_letter(len(de.columns))
    max_row = len(de) + 1
    ws.auto_filter.ref = f"A1:{max_col}{max_row}"
    ws.freeze_panes = "A2"
    ws.sheet_view.rightToLeft = True

    wb.save(out)
    out.seek(0)
    return out.getvalue()


def export_csv(df: pd.DataFrame) -> bytes:
    de  = _safe_export_convert(df).drop(
        columns=[c for c in df.columns if c.startswith("__")],
        errors="ignore"
    )
    out = io.BytesIO()
    out.write("\ufeff".encode("utf-8"))
    de.to_csv(out, index=False, encoding="utf-8")
    out.seek(0)
    return out.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# STYLER - OPTIMIZED FOR LARGE DATASETS
# ═══════════════════════════════════════════════════════════════════════════════
@st.cache_data
def _get_style_config():
    """Cache styling configuration"""
    return {
        "MISS_SET": {"غير محدد", "Not Specified", "غير محدد / Not Specified"},
        "missing_style": "background-color:rgba(248,215,218,0.85);color:#991B1B;font-weight:700;border:1px solid rgba(248,215,218,0.6);",
        "outlier_style": "background-color:rgba(255,243,205,0.85);",
        "median_style": "background-color:rgba(212,237,218,0.85);",
        # NEW: Neon Green for imputed text cells (optimized for dark mode)
        "imputed_text_style": "background-color:rgba(34,197,94,0.25);border:1.5px solid rgba(34,197,94,0.6);color:#86EFAC;font-weight:600;",
    }


def style_imputed_text_cells(df: pd.DataFrame, imputation_mask: pd.DataFrame, preview_only: bool = True, preview_rows: int = 100):
    """
    HIGHLIGHT IMPUTED TEXT CELLS with neon green (dark mode optimized).
    
    This function applies a sleek neon green highlight to cells where:
    - The Dynamic Contextual Text Imputation successfully predicted and filled a missing value.
    
    Args:
        df: Cleaned DataFrame containing the imputed values
        imputation_mask: Boolean mask DataFrame (True where imputation occurred)
        preview_only: If True, only style preview rows (RECOMMENDED for large datasets)
        preview_rows: Number of rows to style (default 100)
    
    Returns:
        Styled DataFrame ready for st.dataframe() display
    
    Example usage in Streamlit app:
        st.subheader("✨ Cleaned Data with AI-Imputed Text Highlighted")
        styled_df = style_imputed_text_cells(
            df=cleaned_df,
            imputation_mask=text_imputation_mask,
            preview_only=True,
            preview_rows=100
        )
        st.dataframe(styled_df, use_container_width=True, height=500)
    """
    try:
        # Use preview for styling (critical for large data!)
        if preview_only and len(df) > preview_rows:
            disp = df.head(preview_rows).copy()
            mask_disp = imputation_mask.head(preview_rows)
        else:
            disp = df.drop(columns=["__outlier__"], errors="ignore")
            mask_disp = imputation_mask
        
        config = _get_style_config()

        def highlight_imputed(data):
            """Apply neon green highlight to imputed cells."""
            # data is a Series for each row
            styles = []
            for col_name, val in data.items():
                try:
                    # Get the mask value for this cell
                    mask_val = mask_disp.loc[data.name, col_name] if col_name in mask_disp.columns else False
                    if mask_val:
                        styles.append(config["imputed_text_style"])
                    else:
                        styles.append("")
                except Exception:
                    styles.append("")
            return styles

        return disp.style.apply(highlight_imputed, axis=1)

    except Exception as e:
        st.warning(f"⚠️ Styling imputed cells: {str(e)[:100]}")
        # Fallback: return unstyled preview
        if preview_only and len(df) > preview_rows:
            return df.head(preview_rows)
        return df.head(preview_rows)


@st.cache_data
def _get_style_config():
    """Cache styling configuration"""
    return {
        "MISS_SET": {"غير محدد", "Not Specified", "غير محدد / Not Specified"},
        "missing_style": "background-color:rgba(248,215,218,0.85);color:#991B1B;font-weight:700;border:1px solid rgba(248,215,218,0.6);",
        "outlier_style": "background-color:rgba(255,243,205,0.85);",
        "median_style": "background-color:rgba(212,237,218,0.85);",
    }

def style_preview(df: pd.DataFrame, preview_only: bool = True, preview_rows: int = 100, median_coords=None, outlier_coords=None, missing_coords=None):
    """
    OPTIMIZED: Style only preview rows to avoid Pandas Styler rendering limits
    
    KEY CHANGE: Works on .head(preview_rows) to avoid "14M+ cells" error
    - Original error: Trying to render only first rows with styling for large datasets
    - Full dataset still processes in background (session state)
    
    Args:
        df: DataFrame (can be 1M+ rows)
        preview_only: If True, only style preview rows (RECOMMENDED)
        preview_rows: Number of rows to style (default 100)
        median_coords: Iterable of (row_index, column_name) tuples to highlight green.
        outlier_coords: Iterable of (row_index, column_name) tuples to highlight yellow.
        missing_coords: Iterable of (row_index, column_name) tuples to highlight red.
    
    Returns:
        Styled DataFrame (safe to render with st.dataframe)
    """
    try:
        # Use preview for styling (critical for large data!)
        if preview_only and len(df) > preview_rows:
            disp = df.head(preview_rows).copy()
        else:
            disp = df.drop(columns=["__outlier__"], errors="ignore")
        
        config = _get_style_config()
        median_coords = set(tuple(c) for c in (median_coords or []))
        outlier_coords = set(tuple(c) for c in (outlier_coords or []))
        missing_coords = set(tuple(c) for c in (missing_coords or []))

        def row_s(row):
            actual_idx = df.index[row.name] if preview_only and len(df) > preview_rows else row.name
            styles = []

            for col, val in row.items():
                coord = (actual_idx, col)
                if coord in median_coords:
                    styles.append(config["median_style"])
                elif coord in outlier_coords:
                    styles.append(config["outlier_style"])
                elif coord in missing_coords:
                    styles.append(config["missing_style"])
                else:
                    styles.append("")
            return styles

        return disp.style.apply(row_s, axis=1)

    except Exception as e:
        st.warning(f"⚠️ Styling preview: {str(e)[:100]}")
        # Fallback: return unstyled preview
        if preview_only and len(df) > preview_rows:
            return df.head(preview_rows)
        return df.head(preview_rows)


# ═══════════════════════════════════════════════════════════════════════════════
# UI HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def render_stepper(cur: int):
    steps = [("1","📤","الرفع"),("2","🔍","التحليل"),
             ("3","⚡","التنظيف"),("4","📥","التصدير")]
    html = '<div class="stepper-wrap">'
    for i,(n,icon,lbl) in enumerate(steps):
        ni = int(n)
        if ni==cur:
            cbg=C['accent']; clbl=C['accent']; ctxt=C['bg']; sh=f"box-shadow:0 0 16px {C['accent']}60;"
        elif ni<cur:
            cbg=C['green']; clbl=C['text2']; ctxt=C['bg']; sh=""
        else:
            cbg=C['bg3']; clbl=C['text3']; ctxt=C['text3']; sh=""
        sym = "✓" if ni<cur else (icon if ni==cur else n)
        html += f"""
        <div class="stepper-step">
            <div class="stepper-circle" style="background:{cbg};color:{ctxt};{sh}">{sym}</div>
            <div class="stepper-label" style="color:{clbl};">{lbl}</div>
        </div>"""
        if i < len(steps) - 1:
            lc = C['green'] if ni < cur else (C['accent'] + "50" if ni == cur else C['border'])
            html += f'<div class="stepper-line" style="background:{lc};"></div>'
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)

def metric_grid(metrics):
    html = '<div class="metric-grid">'
    for (icon,lbl,val,sub) in metrics:
        html += f"""
        <div class="metric-card">
            <div class="metric-icon">{icon}</div>
            <div class="metric-label">{lbl}</div>
            <div class="metric-value">{val}</div>
            <div class="metric-sub">{sub}</div>
        </div>"""
    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)

def stage_header(num, title, desc):
    st.markdown(f"""
    <div class="stage-header">
        <div class="stage-header-num">▸ المرحلة {num}</div>
        <div class="stage-header-title">{title}</div>
        <div class="stage-header-desc">{desc}</div>
    </div>""", unsafe_allow_html=True)

def quality_card(score: float, label="مؤشر جودة البيانات", improvement: float = None):
    clr = C['green'] if score >= 85 else (C['yellow'] if score >= 65 else ("orange" if score >= 40 else C['red']))
    badge = ("ممتاز | Excellent" if score >= 85 else
             ("جيد | Good" if score >= 65 else
              ("مقبول | Fair" if score >= 40 else "ضعيف | Poor")))

    html = f"""
    <div style="padding: 1rem; background: {C['bg1']}; border: 1px solid {C['border']}; border-radius: 18px; width: 100%; min-height: 200px; display: flex; flex-direction: column; align-items: center; justify-content: center;">
        <div style="font-size: 0.78rem; color: {C['text2']}; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; margin-bottom: 1rem; text-align: center;">
            {label}
        </div>
        <div style="font-size: 2.5rem; font-weight: 900; color: {clr}; line-height: 1; margin-bottom: 0.65rem;">{score:.1f}%</div>
        <div style="font-size: 0.95rem; font-weight: 800; color: {clr}; background: rgba(255,255,255,0.06); border: 1px solid rgba(255,255,255,0.12); padding: 0.35rem 0.9rem; border-radius: 999px; text-align: center;">
            {badge}
        </div>
    </div>"""
    st.markdown(html, unsafe_allow_html=True)

def lang_pill(lang: str) -> str:
    labels = {
        "ar":    ("🇸🇦 عربي", "lang-ar"),
        "en":    ("🇬🇧 English", "lang-en"),
        "mixed": ("🌐 مختلط | Mixed", "lang-mixed"),
        "other": ("🔢 أخرى | Other", "lang-en"),
    }
    lbl, cls = labels.get(lang, ("—", "lang-en"))
    return f'<span class="lang-indicator {cls}">{lbl}</span>'

def col_profile_card(col: str, info: dict):
    tp       = info.get("type","text")
    lang     = info.get("language","other")
    pill_cls = {"numeric":"pill-num","text":"pill-text",
                "date":"pill-date","boolean":"pill-bool"}.get(tp,"pill-text")
    pill_lbl = {"numeric":"🔢 رقمي","text":"📝 نصي",
                "date":"📅 تاريخ","boolean":"☑️ ثنائي"}.get(tp,"نصي")
    fp  = info.get("fill_rate", 0)
    np_ = info.get("null_pct",  0)
    bclr= C['green'] if fp>=80 else (C['yellow'] if fp>=50 else C['red'])

    html = f"""
    <div class="col-profile">
        <div class="col-profile-head">
            <span class="col-type-pill {pill_cls}">{pill_lbl}</span>
            <span class="col-name">{col}</span>
            {lang_pill(lang)}
        </div>
        <div class="col-stats">
            <div class="stat-item">
                <span class="stat-label">إجمالي | Total</span>
                <span class="stat-value">{info.get('total',0):,}</span>
            </div>
            <div class="stat-item">
                <span class="stat-label">فريد | Unique</span>
                <span class="stat-value" style="color:{C['cyan']};">
                    {info.get('unique',0):,}
                </span>
            </div>
            <div class="stat-item">
                <span class="stat-label">ناقص | Missing</span>
                <span class="stat-value" style="color:{C['red']};">
                    {info.get('null_count',0):,}
                    <span style="font-size:0.7rem;color:{C['text3']};">({np_}%)</span>
                </span>
            </div>
            <div class="stat-item">
                <span class="stat-label">الإكمال | Fill</span>
                <span class="stat-value" style="color:{bclr};">{fp}%</span>
            </div>"""
    if tp == "numeric" and "mean" in info:
        html += f"""
            <div class="stat-item">
                <span class="stat-label">متوسط | Mean</span>
                <span class="stat-value" style="color:{C['accent']};">{info['mean']:,}</span>
            </div>
            <div class="stat-item">
                <span class="stat-label">أدنى | Min</span>
                <span class="stat-value">{info['min']:,}</span>
            </div>
            <div class="stat-item">
                <span class="stat-label">أعلى | Max</span>
                <span class="stat-value">{info['max']:,}</span>
            </div>
            <div class="stat-item">
                <span class="stat-label">شاذة | Outliers</span>
                <span class="stat-value" style="color:{C['yellow']};">
                    {info.get('outliers',0)}
                </span>
            </div>"""
    html += f"""
        </div>
        <div style="font-size:0.7rem;color:{C['text3']};margin-bottom:0.22rem;">
            معدل الإكمال | Fill Rate
        </div>
        <div class="mini-bar-wrap">
            <div class="mini-bar-fill" style="width:{fp}%;background:{bclr};"></div>
        </div>"""
    if tp == "text" and "top_values" in info:
        chips = "".join(
            f'<span class="val-chip"><strong>{v}</strong> × {str(k)[:22]}</span>'
            for k,v in list(info["top_values"].items())[:5]
        )
        html += f'<div class="top-vals">{chips}</div>'
    pii_type = info.get("pii_type")
    if pii_type:
        html += f"""
        <div style="background:rgba(6,182,212,0.08); border:1px solid rgba(6,182,212,0.2); border-radius:8px; padding:0.45rem 0.7rem; margin-top:0.7rem; font-size:0.76rem; color:{C['cyan']}; font-weight:700; display:flex; align-items:center; gap:0.4rem; direction:rtl;">
            🛡️ تم الكشف عن بيانات حساسة ({pii_type}) في هذا العمود وتم التحقق من سلامة البنية | Sensitive {pii_type} data detected and verified.
        </div>"""
        
    if info.get("recommend_drop"):
        reason_ar = "العمود شبه فارغ (ناقص بنسبة > 90%)" if info.get("is_high_missing") else "العمود ذو تباين منخفض جداً (قيمة واحدة متكررة بنسبة > 98%)"
        reason_en = "Column is > 90% missing" if info.get("is_high_missing") else "Column has zero/low variance (> 98% identical value)"
        html += f"""
        <div style="background:rgba(239,68,68,0.08); border:1px solid rgba(239,68,68,0.2); border-radius:8px; padding:0.45rem 0.7rem; margin-top:0.7rem; font-size:0.76rem; color:#FCA5A5; font-weight:700; display:flex; align-items:center; gap:0.4rem; direction:rtl;">
            ⚠️ توصية بحذف العمود: {reason_ar} | Drop recommended: {reason_en}
        </div>"""

    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)

def alert(msg, kind="info"):
    icons = {"info":"ℹ️","success":"✅","warning":"⚠️","danger":"❌"}
    st.markdown(
        f'<div class="alert alert-{kind}">{icons.get(kind,"ℹ️")} {msg}</div>',
        unsafe_allow_html=True
    )


def log_dataframe_changes(audit_trail, original_df, updated_df, cols, reason):
    if audit_trail is None:
        return
    for idx in original_df.index:
        for col in cols:
            if col in updated_df.columns and original_df.at[idx, col] != updated_df.at[idx, col]:
                audit_trail.log_change(
                    row_idx=idx,
                    col_name=col,
                    old_value=original_df.at[idx, col],
                    new_value=updated_df.at[idx, col],
                    reason=reason,
                )


# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
def render_sidebar():
    with st.sidebar:
        st.markdown(f"""
        <div class="sb-logo">
            <div class="sb-logo-title">✦ DataClean Pro</div>
            <div class="sb-logo-sub">
                بوابة تنظيف البيانات | Data Cleaning Platform
            </div>
        </div>""", unsafe_allow_html=True)

        if st.session_state.df is None:
            st.markdown(f"""
            <div style="padding:2rem 1.2rem;text-align:center;">
                <div style="font-size:2.5rem;margin-bottom:0.8rem;">📂</div>
                <div style="color:{C['text3']};font-size:0.83rem;line-height:1.8;">
                    ابدأ برفع ملف البيانات<br>
                    Start by uploading your data file<br>
                    <span style="font-size:0.75rem;">CSV / Excel</span>
                </div>
            </div>""", unsafe_allow_html=True)
        else:
            df    = st.session_state.df
            clean = df.drop(columns=["__outlier__"], errors="ignore")
            score = _quality_score(clean)
            sc    = C['green'] if score>=85 else (C['yellow'] if score>=65 else C['red'])
            dup   = int(clean.duplicated().sum())
            miss  = int(clean.isnull().sum().sum())

            st.markdown(f"""
            <div class="sb-section">
                <div class="sb-section-title">الملف | File</div>
                <div style="color:{C['accent']};font-weight:700;font-size:0.85rem;
                     word-break:break-all;padding:0.3rem 0;
                     border-bottom:1px solid {C['border']};">
                    📄 {st.session_state.file_name}
                </div>
                <div style="color:{C['text3']};font-size:0.72rem;padding-top:0.3rem;">
                    {st.session_state.file_size/1024:.1f} KB
                </div>
            </div>
            <div class="sb-section">
                <div class="sb-section-title">إحصاءات | Statistics</div>
                <div class="sb-stat-row">
                    <span class="sb-stat-label">الصفوف | Rows</span>
                    <span class="sb-stat-value">{len(clean):,}</span>
                </div>
                <div class="sb-stat-row">
                    <span class="sb-stat-label">الأعمدة | Cols</span>
                    <span class="sb-stat-value">{len(clean.columns):,}</span>
                </div>
                <div class="sb-stat-row">
                    <span class="sb-stat-label">ناقصة | Missing</span>
                    <span class="sb-stat-value" style="color:{C['red']};">{miss:,}</span>
                </div>
                <div class="sb-stat-row">
                    <span class="sb-stat-label">مكررة | Dups</span>
                    <span class="sb-stat-value" style="color:{C['yellow']};">{dup:,}</span>
                </div>
                <div class="sb-stat-row">
                    <span class="sb-stat-label">الجودة | Quality</span>
                    <span class="sb-stat-value" style="color:{sc};">{score}/100</span>
                </div>
            </div>
            """, unsafe_allow_html=True)

            if st.session_state.cleaning_done:
                st.markdown(f"""
                <div style="margin:0.6rem 1.2rem;background:rgba(16,185,129,0.08);
                     border:1px solid rgba(16,185,129,0.2);border-radius:10px;
                     padding:0.7rem;text-align:center;color:{C['green']};
                     font-weight:700;font-size:0.84rem;">
                    ✓ التنظيف مكتمل | Cleaning Done
                </div>""", unsafe_allow_html=True)

            # ═══════════════════════════════════════════════════════════════════════════
            # 🎯 SELECT COLUMNS FOR OUTLIER DETECTION
            # ═══════════════════════════════════════════════════════════════════════════
            st.markdown(f"""
            <div class="sb-section" style="margin-top:1.2rem;">
                <div style="font-size:1rem;font-weight:800;color:{C['text']};
                     margin-bottom:0.3rem;text-align:right;">
                    🎯 اختر الأعمدة | SELECT COLUMNS
                </div>
                <div style="font-size:0.85rem;font-weight:700;color:{C['accent']};
                     margin-bottom:0.6rem;padding-bottom:0.6rem;
                     border-bottom:1px solid {C['border']};text-align:right;">
                    🚨 رصد الشاذات | OUTLIERS
                </div>
            </div>""", unsafe_allow_html=True)
            
            # Get numerical columns and mixed numeric/text columns too
            profile = st.session_state.profile or {}
            numeric_cols = [
                col for col, info in profile.items()
                if info.get("type") == "numeric" and not col.startswith("__")
            ]

            mixed_numeric_cols = []
            if AdvancedCleaner:
                try:
                    mixed_numeric_cols = [
                        col for col in detect_numeric_columns_with_text(df)
                        if not col.startswith("__") and col in df.columns
                    ]
                except Exception:
                    mixed_numeric_cols = []

            numeric_cols = sorted(set(numeric_cols + mixed_numeric_cols))

            # ─── FIX: Fragment for sidebar outlier selector ───
            @st.fragment
            def render_outlier_selector():
                if numeric_cols:
                    def on_outlier_select():
                        st.session_state.selected_outlier_columns = st.session_state.outlier_selector
                    
                    st.multiselect(
                        "اختر أعمدة للكشف عن القيم الشاذة",
                        numeric_cols,
                        default=st.session_state.get("selected_outlier_columns", []),
                        key="outlier_selector",
                        label_visibility="collapsed",
                        help="حدد الأعمدة الرقمية التي تريد تطبيق كشف القيم الشاذة عليها. إذا لم تحدد شيئاً، سيُطبق الكشف على كل الأعمدة الرقمية.",
                        on_change=on_outlier_select
                    )
                    
                    # Show warning if no columns selected when on the cleaning step
                    if len(st.session_state.get("selected_outlier_columns", [])) == 0 and st.session_state.step == 3:
                        st.markdown(f"""
                        <div style="background:rgba(245,158,11,0.08);border:1px solid rgba(245,158,11,0.2);
                             border-radius:8px;padding:0.6rem;margin-top:0.5rem;text-align:right;">
                            <span style="color:{C['yellow']};font-weight:700;font-size:0.85rem;">
                                ⚠️ ⚠️ لم تختر أعمدة للكشف عن الشاذات | No columns selected for outlier detection
                            </span>
                        </div>""", unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div style="background:rgba(59,130,246,0.08);border:1px solid rgba(59,130,246,0.2);
                         border-radius:8px;padding:0.6rem;margin-top:0.5rem;text-align:right;font-size:0.82rem;
                         color:{C['text2']};">
                        ℹ️ لا توجد أعمدة رقمية | No numeric columns available
                    </div>""", unsafe_allow_html=True)
            
            render_outlier_selector()

            st.markdown(f"""
            <div class="sb-section" style="margin-top:1.2rem;">
                <div class="sb-section-title">القدرات | Features</div>
                <div style="font-size:0.77rem;color:{C['text3']};line-height:2.1;">
                    ✦ عربي + إنجليزي | AR + EN<br>
                    ✦ تصحيح التواريخ | Date Fix<br>
                    ✦ تطبيع ذكي | Smart Normalize<br>
                    ✦ استنتاج ثنائي | Bilingual Infer<br>
                    ✦ رصد الشاذات | Outlier Detection<br>
                    ✦ ✨ تنظيف متقدم | Advanced Deep Clean<br>
                    ✦ تصدير احترافي | Pro Export
                </div>
            </div>""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN APP
# ═══════════════════════════════════════════════════════════════════════════════
def apply_sidebar_visibility():
    if not st.session_state.get("sidebar_visible", True):
        st.markdown(
            """
            <style>
            [data-testid="stSidebar"] { display: none !important; }
            section[data-testid="stMain"] { margin-left: 0 !important; }
            .block-container { max-width: 100% !important; padding-left: 3rem !important; padding-right: 3rem !important; }
            </style>
            """,
            unsafe_allow_html=True,
        )


def main():
    inject_css()
    init_state()
    render_sidebar()
    apply_sidebar_visibility()

    sidebar_label = "🧭 إظهار / إخفاء اللوحة الجانبية"
    if st.button(sidebar_label, key="toggle_sidebar"):
        st.session_state.sidebar_visible = not st.session_state.sidebar_visible

    # Hero
    st.markdown(f"""
    <div class="hero">
        <div class="hero-badge">✦ Bilingual Data Cleaning Platform</div>
        <h1 class="hero-title">
            بوابة تنظيف <span>البيانات المتقدمة</span>
        </h1>
        <p class="hero-subtitle">
            منصة احترافية تدعم البيانات
            <span class="lang-badge">🇸🇦 العربية</span>
            و
            <span class="lang-badge">🇬🇧 English</span>
            بشكل كامل — 10 خطوات ذكية
        </p>
    </div>""", unsafe_allow_html=True)

    render_stepper(st.session_state.step)
    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # STEP 1 — UPLOAD
    # ══════════════════════════════════════════════════════════════════════════
    if st.session_state.step == 1:
        stage_header(
            "الأولى | First", "📤 رفع البيانات | Upload Data",
            "أرسل ملف CSV أو Excel بأي لغة — عربي، إنجليزي، أو مختلط. "
            "النظام يكتشف اللغة تلقائياً لكل عمود ويطبّق معالجة دقيقة ومتكاملة. | "
            "Upload CSV or Excel in Arabic, English, or mixed. "
            "The system auto-detects the language for each column and applies precise processing."
        )

        up = st.file_uploader(
            "حدد ملف CSV أو Excel | Select CSV / Excel",
            type=["csv","xlsx","xls"],
        )

        if up:
            try:
                with st.spinner("⏳ جاري القراءة والتحليل | Reading & analyzing..."):
                    df = read_file(up)
                    
                    # ─── OPTIMIZATION: Memory Optimization ───
                    df_optimized, mem_stats = optimize_dtypes(df)
                    df = df_optimized
                    
                    # Build profile (expensive, but needed once)
                    profile = build_profile(df)
                    score = _quality_score(df)

                    # كشف لغة كل عمود
                    lang_summary = {}
                    for col in df.columns:
                        if _detect_type(df[col]) == "text":
                            lang_summary[col] = detect_column_language(df[col])

                st.session_state.update({
                    "df": df.copy(), "df_original": df.copy(),
                    "file_name": up.name, "file_size": up.size,
                    "step": 2, "profile": profile,
                    "quality_before": score,
                    "detected_langs": lang_summary,
                    "memory_stats": mem_stats,  # Track optimization
                })
                
                # Show memory optimization results
                if mem_stats.get("saved_mb", 0) > 0:
                    st.success(
                        f"✓ Memory optimized! Saved **{mem_stats['saved_mb']} MB** "
                        f"({mem_stats['reduction_pct']}% reduction)"
                    )
                
                alert(
                    f"تم التحميل | Uploaded — {len(df):,} صفوف | rows · "
                    f"{len(df.columns):,} أعمدة | cols · "
                    f"جودة | Quality: {score}/100",
                    "success"
                )
                st.rerun()
            except Exception as e:
                alert(f"خطأ | Error: {e}", "danger")

        # Features
        feats = [
            ("🌐","ثنائي اللغة | Bilingual",
             "يعالج العربي والإنجليزي والمختلط في نفس الملف | "
             "Handles Arabic, English & mixed data"),
            ("📅","تصحيح التواريخ | Date Fix",
             "يحول كل صيغ التواريخ بما فيها أرقام Excel | "
             "Converts all date formats including Excel serials"),
            ("🧠","استنتاج ذكي | Smart Inference",
             "يملأ الفراغات بناءً على الكلمات المفتاحية بالعربي والإنجليزي | "
             "Fills gaps using AR+EN keywords"),
            ("🚨","رصد الشاذات | Outliers",
             "خوارزمية IQR لاكتشاف القيم غير الطبيعية | "
             "IQR algorithm for anomaly detection"),
            ("🔠","تطبيع ذكي | Smart Normalize",
             "تطبيع مخصص حسب لغة كل عمود | "
             "Custom normalization per column language"),
            ("📊","تحليل شامل | Full Analysis",
             "إحصاءات وكشف لغة لكل عمود | "
             "Stats & language detection per column"),
            ("📥","تصدير احترافي | Pro Export",
             "Excel مع محاذاة صحيحة للعربي والإنجليزي | "
             "Excel with correct AR/EN alignment"),
            ("⚡","10 خطوات | 10 Steps",
             "خط معالجة متكامل في ثوانٍ | "
             "Complete pipeline in seconds"),
        ]
        html = '<div class="feature-grid">'
        for icon,title,desc in feats:
            html += f"""
            <div class="feature-card">
                <span class="feature-icon">{icon}</span>
                <div class="feature-title">{title}</div>
                <div class="feature-desc">{desc}</div>
            </div>"""
        html += '</div>'
        st.markdown(html, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # STEP 2 — ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    elif st.session_state.step == 2:
        df      = st.session_state.df
        profile = st.session_state.profile or build_profile(df)
        score   = _quality_score(df)

        stage_header(
            "الثانية | Second",
            "🔍 تحليل البيانات | Data Analysis",
            "تحليل شامل لكل عمود مع كشف اللغة التلقائي — "
            "عربي 🇸🇦، إنجليزي 🇬🇧، أو مختلط 🌐 | "
            "Full column analysis with auto language detection — AR, EN, or Mixed."
        )

        # Language Summary
        lang_counts = {"ar":0,"en":0,"mixed":0,"other":0}
        for info in profile.values():
            l = info.get("language","other")
            lang_counts[l] = lang_counts.get(l,0)+1

        qc, mc = st.columns([1,3])
        with qc:
            quality_card(score)
        with mc:
            miss  = int(df.isnull().sum().sum())
            dups  = int(df.duplicated().sum())
            dates = sum(1 for c in df.columns if _is_date_column(df[c]))
            metric_grid([
                ("📊","الصفوف | Rows",   f"{len(df):,}",   ""),
                ("📋","الأعمدة | Cols",  f"{len(df.columns):,}", ""),
                ("❓","ناقصة | Missing", f"{miss:,}",
                 f"{round(miss/max(df.size,1)*100,1)}%"),
                ("♻️","مكررة | Dups",   f"{dups:,}",       ""),
                ("🇸🇦","عربي | Arabic",  f"{lang_counts['ar']}",  "عمود | col"),
                ("🇬🇧","إنجليزي | EN",  f"{lang_counts['en']}",  "عمود | col"),
                ("🌐","مختلط | Mixed",   f"{lang_counts['mixed']}","عمود | col"),
                ("📅","تاريخ | Date",    f"{dates}",         "عمود | col"),
            ])

        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
        
        # Scan for PII columns and drop recommendations in Step 2
        pii_cols = [col for col, info in profile.items() if info.get("pii_type")]
        drop_cols = [col for col, info in profile.items() if info.get("recommend_drop")]
        
        if pii_cols:
            pii_details = ", ".join(f"**{col}** ({profile[col]['pii_type']})" for col in pii_cols)
            alert(f"🛡️ **الكشف عن بيانات حساسة (PII Auto-Detection):** تم التعرف على بيانات حساسة في الأعمدة التالية: {pii_details} وتم فحص تركيبها بنجاح. | **PII Auto-Detection:** Sensitive data identified in: {pii_details}. Structure verified successfully.", "success")
            
        if drop_cols:
            drop_details = []
            for col in drop_cols:
                reason = "شبه فارغ" if profile[col].get("is_high_missing") else "تباين منخفض جداً"
                reason_en = "high-missing" if profile[col].get("is_high_missing") else "low-variance"
                drop_details.append(f"**{col}** ({reason} / {reason_en})")
            drop_details_str = ", ".join(drop_details)
            alert(f"⚠️ **توصية بتنظيف الهيكل (Structural Recommendation):** يوصى بحذف الأعمدة التالية لأنها عديمة الفائدة إحصائياً وسيقوم المحرك بحذفها تلقائياً: {drop_details_str} | **Structure Alert:** Columns recommended for dropping (dropped automatically during cleaning): {drop_details_str}.", "warning")

        t1,t2,t3 = st.tabs([
            "👀  معاينة | Preview",
            "📊  تحليل الأعمدة | Columns",
            "🔴  القيم الناقصة | Missing",
        ])

        with t1:
            alert(
                "الخلايا الحمراء = قيم ناقصة | Red cells = missing values",
                "info"
            )
            # ─── OPTIMIZATION: Safe dataframe display for large datasets ───
            # Shows preview only (100 rows), but all data is processed
            preview_rows = min(100, len(df))
            if len(df) > preview_rows:
                st.info(
                    f"👀 **Preview Mode**: Showing first {preview_rows:,} of {len(df):,} rows  \n"
                    f"💾 All {len(df):,} rows are processed in background  \n"
                    f"📥 Download the cleaned file below to get complete results"
                )
            
            st.dataframe(
                style_preview(
                    df,
                    preview_only=True,
                    preview_rows=preview_rows,
                    median_coords=st.session_state.get("median_imputed_cells", []),
                    outlier_coords=st.session_state.get("yellow_coordinates", []),
                    missing_coords=st.session_state.get("red_coordinates", []),
                ),
                use_container_width=True,
                height=420
            )

        with t2:
            for col,info in profile.items():
                col_profile_card(col, info)

        with t3:
            miss_df = (
                df.isnull().sum()
                .reset_index()
                .rename(columns={"index":"العمود | Column",0:"الناقصة | Missing"})
            )
            miss_df = miss_df[miss_df["الناقصة | Missing"]>0].sort_values(
                "الناقصة | Missing", ascending=False
            )
            if miss_df.empty:
                alert("لا توجد قيم ناقصة! | No missing values!", "success")
            else:
                for _,row in miss_df.iterrows():
                    pct  = round(row["الناقصة | Missing"]/len(df)*100, 1)
                    clr  = C['red'] if pct>30 else (C['yellow'] if pct>10 else C['accent'])
                    st.markdown(f"""
                    <div class="miss-row">
                        <div class="miss-row-head">
                            <span class="miss-col-name">{row['العمود | Column']}</span>
                            <span class="miss-pct" style="color:{clr};">
                                {int(row['الناقصة | Missing']):,} ({pct}%)
                            </span>
                        </div>
                        <div class="miss-bar-wrap">
                            <div class="miss-bar-fill"
                                 style="width:{pct}%;background:{clr};"></div>
                        </div>
                    </div>""", unsafe_allow_html=True)

        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
        cb, cn = st.columns(2)
        with cb:
            if st.button("⬅️ العودة | Back"):
                st.session_state.df   = None
                st.session_state.step = 1
                st.rerun()
        with cn:
            if st.button("⚡ الانتقال للتنظيف | Go to Cleaning"):
                st.session_state.step = 3
                st.rerun()

    # ══════════════════════════════════════════════════════════════════════════
    # STEP 3 — CLEANING
    # ══════════════════════════════════════════════════════════════════════════
    elif st.session_state.step == 3:
        stage_header(
            "الثالثة | Third",
            "⚡ محرك التنظيف | Cleaning Engine",
            "10 خطوات احترافية تعمل بذكاء على البيانات العربية والإنجليزية معاً | "
            "10 professional steps working intelligently on Arabic & English data."
        )

        steps_data = [
            ("إزالة الأعمدة المكررة", "Remove Duplicate Columns",
             "كشف بالـ Hash لتحديد الأعمدة المتطابقة", "Hash-based detection of identical columns", "🔁", ""),
            ("تنظيف المسافات", "Whitespace Cleaning",
             "حذف مسافات الطرفين والمسافات الزائدة", "Leading/trailing & extra spaces removal", "✂️", ""),
            ("تصحيح التواريخ", "Date Fix",
             "تحويل كافة صيغ التواريخ والأرقام المتسلسلة إلى YYYY-MM-DD", "Convert all date formats & serials to YYYY-MM-DD", "📅", "new"),
            ("تطبيع ثنائي اللغة", "Bilingual Normalization",
             "معالجة متطورة للنصوص العربية والإنجليزية", "Advanced Arabic NLP & English text normalization", "🔠", "bilingual"),
            ("استنتاج متقدم ثنائي", "Advanced Bilingual Inference",
             "استنتاج قيم التصنيفات باستخدام كلمات مفتاحية", "Category filling using bilingual keywords", "🧠", "advanced"),
            ("تنظيف متقدم عميق", "Advanced Deep Cleaning",
             "تصحيح الأخطاء الإملائية واستنتاج السياق", "Fix typos & infer values using context", "✨", "advanced"),
            ("توحيد الأنواع", "Type Standardization",
             "تحويل النصوص إلى أرقام أو قيم منطقية عند الحاجة", "Standardize string to numeric or boolean values", "🔢", "new"),
            ("رصد الشاذات", "Outlier Detection",
             "تطبيق خوارزمية IQR على الأعمدة الرقمية للكشف عن القيم الشاذة", "IQR anomaly detection on numeric columns", "🚨", "new"),
            ("حذف المكررات", "Remove Duplicates",
             "إزالة الصفوف المتكررة بالكامل والإبقاء على الأول", "Remove identical rows, preserving the first", "🗑️", ""),
            ("ملء ذكي", "Smart Fill",
             "ملء القيم المفقودة بـ 'غير محدد' حسب لغة العمود", "Fill missing values with 'Not Specified' by column language", "✏️", "bilingual"),
            ("إعادة التقييم", "Re-evaluation",
             "حساب درجة جودة البيانات قبل وبعد عملية التنظيف", "Compute quality score before and after cleaning", "📈", ""),
        ]
        badge_map = {
            "new":       '<span class="step-badge badge-new">✨ جديد</span>',
            "advanced":  '<span class="step-badge badge-advanced">🔥 متقدم</span>',
            "bilingual": '<span class="step-badge badge-bilingual">🌐 ثنائي</span>',
            "":          "",
        }
        for i,(title_ar, title_en, desc_ar, desc_en, icon, badge) in enumerate(steps_data,1):
            st.markdown(f"""
            <div class="step-item" style="border: 1px solid {C['border']}; border-radius: 12px; padding: 1.2rem; background: {C['bg2']}; margin-bottom: 0.8rem; display: flex; align-items: flex-start; gap: 1rem; position: relative;">
                <div class="step-num" style="background: {C['bg3']}; color: {C['accent']}; font-weight: 800; width: 30px; height: 30px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 0.9rem; border: 1px solid {C['border']}; flex-shrink: 0;">{i}</div>
                <div class="step-icon" style="font-size: 1.3rem; flex-shrink: 0; margin-top: 0.1rem;">{icon}</div>
                <div class="step-content" style="flex-grow: 1; width: 100%;">
                    <div class="step-title-row" style="display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 0.5rem; width: 100%;">
                        <span style="font-weight: 700; color: {C['text']}; font-size: 0.95rem; direction: rtl; text-align: right;">{title_ar}</span>
                        <span style="font-size: 0.8rem; color: {C['text3']}; font-weight: 500; direction: ltr; text-align: left;" class="en-text">{title_en}</span>
                    </div>
                    <div class="step-desc-row" style="display: flex; justify-content: space-between; align-items: flex-start; flex-wrap: wrap; gap: 0.5rem; margin-top: 0.4rem; border-top: 1px solid {C['border']}40; padding-top: 0.4rem;">
                        <span style="font-size: 0.8rem; color: {C['text2']}; direction: rtl; text-align: right; max-width: 50%;">{desc_ar}</span>
                        <span style="font-size: 0.75rem; color: {C['text3']}; direction: ltr; text-align: left; max-width: 50%;" class="en-text">{desc_en}</span>
                    </div>
                </div>
                <div style="position: absolute; top: -10px; left: 10px;">
                    {badge_map.get(badge,"")}
                </div>
            </div>""", unsafe_allow_html=True)

        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

        # ─── ADVANCED OPTIONS ───
        if AdvancedCleaner:
            with st.expander("✨ خيارات متقدمة | Advanced Options", expanded=False):
                st.write("**تفعيل الميزات الاحترافية | Enable Professional Features:**")
                
                col1, col2 = st.columns(2)
                with col1:
                    adv_opts = st.session_state.get("advanced_options", {})
                    
                    st.session_state.advanced_options["clean_phones"] = st.checkbox(
                        "☎️ توحيد أرقام التليفونات",
                        value=adv_opts.get("clean_phones", False),
                        help="تحويل جميع صيغ التليفونات إلى صيغة موحدة (+201XXXXXXXXX)"
                    )
                    
                    st.session_state.advanced_options["clean_numeric"] = st.checkbox(
                        "💰 تنظيف الأرقام المخلوطة",
                        value=adv_opts.get("clean_numeric", False),
                        help="إزالة الرموز والنصوص من الأعمدة الرقمية (1500 ج.م → 1500)"
                    )
                    
                    st.session_state.advanced_options["fuzzy_match"] = st.checkbox(
                        "🎯 الدمج الذكي للمتشابهات",
                        value=adv_opts.get("fuzzy_match", False),
                        help="دمج القيم المتشابهة جداً (القاهرة، القاهره → القاهرة)"
                    )
                
                with col2:
                    st.session_state.advanced_options["fix_dates"] = st.checkbox(
                        "📅 التحقق المتقدم من التواريخ",
                        value=adv_opts.get("fix_dates", False),
                        help="التحقق من صحة التواريخ والتواريخ المستقبلية المستحيلة"
                    )
                    
                    st.session_state.advanced_options["optimize_memory"] = st.checkbox(
                        "⚙️ تحسين استهلاك الذاكرة",
                        value=adv_opts.get("optimize_memory", False),
                        help="تقليل حجم البيانات في الذاكرة (float64→float32, إلخ)"
                    )
                    st.session_state.advanced_options["auto_drop_low_variance"] = st.checkbox(
                        "🧹 حذف الأعمدة الفارغة أو منخفضة التباين تلقائياً",
                        value=adv_opts.get("auto_drop_low_variance", False),
                        help="إزالة الأعمدة التي تحتوي على أكثر من 90% قيم ناقصة أو أكثر من 98% قيمة مكررة"
                    )
                
                st.info(
                    "💡 **نصيحة | Tip:** استخدم جميع الخيارات المتقدمة للحصول على أفضل جودة تنظيف!\n"
                    "Use all advanced options for the best cleaning quality!"
                )

        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
        cb, cr = st.columns(2)
        with cb:
            if st.button("⬅️ العودة | Back"):
                st.session_state.step = 2
                st.rerun()
        with cr:
            if st.button("⚡ تشغيل محرك التنظيف | Run Cleaning Engine"):
                prog = st.progress(0)
                box  = st.empty()
                lbls = [
                    "إزالة الأعمدة المكررة | Removing duplicate columns...",
                    "تنظيف المسافات | Cleaning whitespace...",
                    "تصحيح التواريخ | Fixing dates...",
                    "تطبيع النصوص AR+EN | Normalizing AR+EN text...",
                    "الاستنتاج الذكي | Running smart inference...",
                    "✨ التنظيف العميق | Deep cleaning...",
                    "توحيد الأنواع | Standardizing types...",
                    "رصد الشاذات | Detecting outliers...",
                    "حذف المكررات | Removing duplicates...",
                    "ملء الفراغات | Filling missing values...",
                    "إعادة حساب الجودة | Recalculating quality...",
                ]
                for si,lbl in enumerate(lbls):
                    box.markdown(
                        f'<div class="alert alert-info">⏳ {lbl}</div>',
                        unsafe_allow_html=True
                    )
                    prog.progress(min(int((si + 1) * 10), 100))  
                run_cleaning()
                prog.progress(100)
                box.markdown(
                    f'<div class="alert alert-success">'
                    f'✅ اكتمل التنظيف | Cleaning Complete!</div>',
                    unsafe_allow_html=True
                )
                st.rerun()

    # ══════════════════════════════════════════════════════════════════════════
    # STEP 4 — EXPORT
    # ══════════════════════════════════════════════════════════════════════════
    elif st.session_state.step == 4:
        df = st.session_state.df

        stage_header(
            "الرابعة | Fourth",
            "📥 التصدير والتقرير | Export & Report",
            "بياناتك نظيفة ومحسّنة. راجع التقرير وصدّر الملف | "
            "Your data is clean & optimized. Review the report and export."
        )

        # Impact
        st.markdown(f"""
        <div style="display: flex; justify-content: space-between; align-items: center; margin: 1rem 0 0.7rem; flex-wrap: wrap; gap: 0.5rem; width: 100%;">
            <h3 style="font-size:1.1rem;font-weight:800; color:{C['text']}; margin: 0;">
                📈 تقرير التأثير | Impact Report
            </h3>
            <span style="font-size: 0.8rem; font-weight: 700; color: {C['green']}; background: rgba(16,185,129,0.1); border: 1px solid rgba(16,185,129,0.3); padding: 0.3rem 0.8rem; border-radius: 50px;" class="en-text">
                ⏱️ تم التنظيف في {st.session_state.get('cleaning_time', 1.25):.2f} ثانية | Cleaned in {st.session_state.get('cleaning_time', 1.25):.2f}s
            </span>
        </div>
        """, unsafe_allow_html=True)

        impacts = [
            ("🗑️","صفوف محذوفة | Rows Removed",
             f"{st.session_state.rows_deleted:,}","مكررة | duplicates"),
            ("📋","أعمدة محذوفة | Cols Removed",
             f"{st.session_state.dup_cols_removed:,}","مكررة | duplicates"),
            ("✏️","خلايا مُعالجة | Cells Fixed",
             f"{st.session_state.cells_fixed:,}","استنتاج | inference"),
            ("📅","تواريخ مُصلحة | Dates Fixed",
             f"{st.session_state.dates_fixed:,}","YYYY-MM-DD"),
            ("🚨","قيم شاذة | Outliers",
             f"{st.session_state.outliers_flagged:,}","مُعلَّمة | flagged"),
            ("🔢","أعمدة محوّلة | Cols Converted",
             f"{st.session_state.types_fixed:,}","نوع | type"),
            ("📊","الصفوف | Rows",
             f"{st.session_state.rows_before:,}→{st.session_state.rows_after:,}",
             "قبل→بعد | before→after"),
        ]
        html = '<div class="impact-grid">'
        for icon,lbl,val,sub in impacts:
            html += f"""
            <div class="impact-card">
                <div class="impact-icon">{icon}</div>
                <div class="impact-val">{val}</div>
                <div class="impact-lbl">{lbl}</div>
                <div class="impact-sub">{sub}</div>
            </div>"""
        html += '</div>'
        st.markdown(html, unsafe_allow_html=True)

        # Quality
        qa_c, qb_c = st.columns(2)
        with qa_c:
            quality_card(st.session_state.quality_before,
                         "الجودة قبل | Quality Before")
        with qb_c:
            improvement = st.session_state.quality_after - st.session_state.quality_before
            quality_card(st.session_state.quality_after,
                         "الجودة بعد | Quality After",
                         improvement=improvement)

        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

        # Tabs
        audit_trail = st.session_state.get("audit_trail")
        t1,t2,t3,t4,t5 = st.tabs([
            "📊  البيانات | Data",
            "📋  الأعمدة | Columns",
            "🚨  الشاذات | Outliers",
            "📝  السجل | Log",
            "📋  سجل التغييرات | Audit Trail",
        ])
        
        with t1:
            alert(
                "🔴 أحمر | Red = غير محدد / Not Specified  |  "
                "🟡 أصفر | Yellow = قيمة شاذة | Outlier  |  "
                "✨ أخضر نيون | Neon Green = نص مستنتج بذكاء | AI-Imputed Text  |  "
                "📅 تواريخ | Dates = YYYY-MM-DD",
                "info"
            )
            st.dataframe(
                style_preview(
                    df.head(50),
                    preview_only=True,
                    preview_rows=50,
                    median_coords=st.session_state.get("median_imputed_cells", []),
                    outlier_coords=st.session_state.get("yellow_coordinates", []),
                    missing_coords=st.session_state.get("red_coordinates", []),
                ),
                use_container_width=True, height=420
            )
            
            # NEW: Display imputed text cells with neon green highlighting
            text_mask = st.session_state.get("text_imputation_mask")
            if text_mask is not None and (text_mask.sum().sum() > 0):
                st.markdown("---")
                st.subheader("✨ النصوص المستنتجة بذكاء | AI-Imputed Text Cells")
                st.markdown(
                    f"<div style='font-size:0.9rem;color:#86EFAC;font-weight:600;margin-bottom:1rem;'>"
                    f"🎯 تم استنتاج واستكمال **{int(text_mask.sum().sum())}** خلية نصية بناءً على السياق والعلاقات في البيانات | "
                    f"**{int(text_mask.sum().sum())}** text cells were intelligently imputed based on context analysis."
                    f"</div>",
                    unsafe_allow_html=True
                )
                try:
                    styled_imputed = style_imputed_text_cells(
                        df=df.head(50),
                        imputation_mask=text_mask.head(50),
                        preview_only=True,
                        preview_rows=50
                    )
                    st.dataframe(styled_imputed, use_container_width=True, height=300)
                except Exception as e:
                    st.warning(f"Could not display imputed cells: {str(e)[:50]}")

        with t2:
            pa = build_profile(df)
            for col,info in pa.items():
                col_profile_card(col, info)

        with t3:
            if "__outlier__" in df.columns:
                od = df[df["__outlier__"]==True].drop(
                    columns=["__outlier__"], errors="ignore"
                )
                if od.empty:
                    alert("لم يُرصد أي قيم شاذة! | No outliers detected!", "success")
                else:
                    alert(
                        f"{len(od):,} صف يحتوي على قيم شاذة | "
                        f"rows with outliers (IQR method)",
                        "warning"
                    )
                    st.dataframe(od.head(30), use_container_width=True, height=350)
            else:
                alert("لا توجد بيانات شاذة | No outlier data", "info")

        with t4:
            for line in st.session_state.cleaning_log:
                if "|" in line:
                    parts = line.split("|")
                    ar_part = parts[0].strip()
                    en_part = parts[1].strip()
                    st.markdown(
                        f'<div class="log-item" style="display: flex; justify-content: space-between; align-items: center; width: 100%; direction: rtl;">'
                        f'  <div style="display: flex; align-items: center; gap: 0.6rem; text-align: right;">'
                        f'      <span class="log-check">✓</span>'
                        f'      <span style="font-weight: 700; color: {C["text"]};">{ar_part}</span>'
                        f'  </div>'
                        f'  <span class="en-text" style="font-size: 0.76rem; color: {C["text3"]}; text-align: left; direction: ltr;">{en_part}</span>'
                        f'</div>',
                        unsafe_allow_html=True
                    )
                else:
                    st.markdown(
                        f'<div class="log-item" style="display: flex; align-items: center; gap: 0.6rem; direction: rtl; text-align: right;">'
                        f'  <span class="log-check">✓</span>'
                        f'  <span style="font-weight: 700; color: {C["text"]};">{line}</span>'
                        f'</div>',
                        unsafe_allow_html=True
                    )
        
        # Audit Trail Tab
        with t5:
            if audit_trail is not None and not audit_trail.empty:
                st.subheader("📋 سجل التغييرات | Change Log")
                alert(
                    "جميع التغييرات التي تمت على البيانات، الصف رقم، العمود، القيمة القديمة والجديدة | "
                    "All changes made to the data, row number, column, old and new values",
                    "info"
                )
                st.dataframe(audit_trail, use_container_width=True, height=400)
                
                # Download Audit Trail
                audit_csv = audit_trail.to_csv(index=False, encoding='utf-8-sig')
                with io.BytesIO() as excel_buffer:
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        audit_trail.to_excel(writer, index=False)
                    excel_data = excel_buffer.getvalue()

                col_csv, col_xlsx = st.columns(2)
                with col_csv:
                    st.download_button(
                        label="📥 تنزيل سجل التغييرات CSV | Download Change Log CSV",
                        data=audit_csv,
                        file_name=f"audit_trail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                    )
                with col_xlsx:
                    st.download_button(
                        label="📥 تنزيل سجل التغييرات Excel | Download Change Log Excel",
                        data=excel_data,
                        file_name=f"audit_trail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            else:
                alert("لا يوجد سجل تغييرات حالياً. | No audit trail available yet.", "info")

        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

        # Export
        st.markdown(f"""
        <h3 style="font-size:1.1rem;font-weight:800;color:{C['text']};
             margin:0.5rem 0 0.8rem;text-align:center;">
            📂 اختر تنسيق التصدير | Choose Export Format
        </h3>""", unsafe_allow_html=True)

        fmt = st.radio(
            "", ["📊 Excel (.xlsx)","📄 CSV (.csv)"],
            horizontal=True, label_visibility="collapsed",
        )

        if "Excel" in fmt:
            alert(
                "Excel احترافي | Professional Excel: "
                "محاذاة AR يمين + EN يسار ✓ | AR right-align + EN left-align ✓ | "
                "تواريخ YYYY-MM-DD ✓ | "
                "غير محدد / Not Specified باللون الأحمر ✓ | "
                "فلتر + تجميد ✓",
                "info"
            )
            with st.spinner("⏳ جاري الإنشاء | Creating Excel..."):
                data  = export_excel(df)
            fname = "cleaned_data.xlsx"
            mime  = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            lbl   = "⬇️ تنزيل Excel | Download Excel"
        else:
            alert(
                "CSV مع UTF-8 BOM لدعم العربية والإنجليزية | "
                "CSV with UTF-8 BOM for full AR+EN support ✓",
                "info"
            )
            with st.spinner("⏳ جاري الإنشاء | Creating CSV..."):
                data  = export_csv(df)
            fname = "cleaned_data.csv"
            mime  = "text/csv"
            lbl   = "⬇️ تنزيل CSV | Download CSV"

        cd, cr = st.columns(2)
        with cd:
            st.download_button(lbl, data, fname, mime, key="dl_btn")
        with cr:
            if st.button("🔄 بدء جديد | New Session"):
                for k in list(st.session_state.keys()):
                    del st.session_state[k]
                st.rerun()


if __name__ == "__main__":
    main()