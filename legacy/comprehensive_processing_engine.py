"""
Comprehensive Processing Engine

This file contains a production-quality, high-performance data processing engine
for robust validation, numeric/date normalization, unexpected shift detection,
audit reporting, and optional Excel export.

Version: 2.0
Author: Data Quality Team
"""

import difflib
import json
import logging
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

MISSING_VALUES = {
    "", "unknown", "na", "n/a", "none", "null", "nan", "nat",
    "#n/a", "-", "--", "?", "missing", "not available",
    "غير معروف", "غير معرف", "غير محدد", "لا يوجد",
}

__all__ = ["ComprehensiveProcessingEngine", "ProcessingConfig", "ProcessingResult"]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

NUMERIC_FIELD_KEYWORDS = [
    "price", "amount", "total", "cost", "fee", "tax", "value",
    "balance", "sum", "subtotal", "net", "gross", "paid", "charge",
    "invoice", "discount", "salary", "commission", "budget",
]
ARABIC_NUMERIC_FIELD_KEYWORDS = [
    "سعر", "مبلغ", "تكلفة", "قيمة", "إجمالي", "المجموع", "رصيد",
    "دفع", "فاتورة", "ضريبة", "رسوم", "صاف", "دفعة",
]

MISSING_VALUES = {
    "", "unknown", "na", "n/a", "none", "null", "nan", "nat",
    "#n/a", "-", "--", "?", "missing", "not available",
    "undefined", "not specified", "not provided", "not found",
    "not applicable", "no data", "no value", "empty",
    "غير معروف", "غير معرف", "غير محدد", "لا يوجد",
    "مفقود", "غير متاح", "غير موجود", "غير مدرج",
    "غير معرّف", "لا شيء", "فارغ", "غير مكتمل",
}


def _is_financial_column_name(col_name: str) -> bool:
    lower = str(col_name).lower()
    return any(keyword in lower for keyword in NUMERIC_FIELD_KEYWORDS) or any(keyword in lower for keyword in ARABIC_NUMERIC_FIELD_KEYWORDS)


@dataclass
class ProcessingConfig:
    """Configuration object for processing behavior."""

    date_columns: List[str] = field(default_factory=list)
    numeric_columns: List[str] = field(default_factory=list)
    notes_column: str = "Notes"
    audit_sheet_name: str = "Audit Report"
    include_audit_sheet: bool = True
    category_threshold: float = 0.20
    optimize_memory: bool = True


@dataclass
class ProcessingResult:
    """Result container returned by the engine pipeline."""

    dataframe: pd.DataFrame
    audit_report: Dict[str, Any]
    duration_seconds: float
    status: str = "success"


class ComprehensiveProcessingEngine:
    """A scalable pipeline for professional data cleaning and auditability."""

    def __init__(self, df: pd.DataFrame, config: Optional[ProcessingConfig] = None):
        self.df = df.copy(deep=True)
        self.original_df = df.copy(deep=True)
        self.config = config or ProcessingConfig()
        self._ensure_notes_column()
        self._infer_columns_if_missing()
        self._init_metrics()
        logger.info("Initialized ComprehensiveProcessingEngine")

    @staticmethod
    def load_dataframe(file_path: str, encodings: Optional[List[str]] = None) -> pd.DataFrame:
        """Read CSV with encoding fallbacks for robust ingestion."""
        encodings = encodings or [
            "utf-8",
            "utf-8-sig",
            "cp1256",
            "windows-1256",
            "iso-8859-6",
            "latin-1",
        ]

        for encoding in encodings:
            try:
                return pd.read_csv(file_path, encoding=encoding)
            except Exception:
                continue

        return pd.read_csv(file_path, encoding="utf-8", errors="replace")

    def _init_metrics(self) -> None:
        self.metrics = {
            "total_rows": len(self.df),
            "total_columns": len(self.df.columns),
            "total_cells": len(self.df) * len(self.df.columns),
            "cells_checked": 0,
            "cells_modified": 0,
            "cells_moved_to_notes": 0,
            "type_conversions": 0,
            "date_sterilized": 0,
            "null_replacements": 0,
            "memory_before_bytes": self.df.memory_usage(deep=True).sum(),
            "memory_after_bytes": None,
        }
        self.audit_log: List[Dict[str, Any]] = []
        self.type_violations: List[Dict[str, Any]] = []
        self.data_shift_log: List[Dict[str, Any]] = []
        self.green_coordinates: List[Tuple[int, str]] = []
        self.yellow_coordinates: List[Tuple[int, str]] = []
        self.red_coordinates: List[Tuple[int, str]] = []
        self.numeric_cols: List[str] = []
        self.text_cols: List[str] = []

    def _ensure_notes_column(self) -> None:
        if self.config.notes_column not in self.df.columns:
            self.df[self.config.notes_column] = pd.Series([""] * len(self.df), dtype="string")
        else:
            self.df[self.config.notes_column] = self.df[self.config.notes_column].astype("string").fillna("")

    def _infer_columns_if_missing(self) -> None:
        if not self.config.numeric_columns:
            self.config.numeric_columns = self._infer_numeric_columns()
        if not self.config.date_columns:
            self.config.date_columns = self._infer_date_columns()

    def _infer_numeric_columns(self) -> List[str]:
        numeric_columns: List[str] = []
        for col in self.df.columns:
            series = self.df[col]
            if pd.api.types.is_numeric_dtype(series):
                numeric_columns.append(col)
                continue
            sample = series.dropna().astype(str).head(50)
            if sample.empty:
                continue
            cleaned = sample.str.replace(r"[,\s]", "", regex=True)
            converted = pd.to_numeric(cleaned, errors="coerce")
            if converted.notna().mean() >= 0.85:
                numeric_columns.append(col)
        return numeric_columns

    def _infer_date_columns(self) -> List[str]:
        date_columns: List[str] = []
        for col in self.df.columns:
            sample = self.df[col].dropna().astype(str).head(50)
            if sample.empty:
                continue
            parsed = pd.to_datetime(sample, errors="coerce", infer_datetime_format=True)
            if parsed.notna().mean() >= 0.75:
                date_columns.append(col)
        return date_columns

    def _force_strict_numeric_keywords(self) -> None:
        for col in self.df.columns:
            if _is_financial_column_name(col):
                coerced = pd.to_numeric(self.df[col], errors="ignore")
                if coerced.dtype != object:
                    self.df[col] = coerced

    def _sync_column_sets(self) -> None:
        self.numeric_cols = self.df.select_dtypes(include=["number"]).columns.tolist()
        self.text_cols = self.df.select_dtypes(include=["object", "string", "category"]).columns.tolist()

    def _drop_blank_strings_for_text(self, raw: pd.Series) -> pd.Series:
        if pd.api.types.is_string_dtype(raw) or raw.dtype == object:
            return raw.where(~raw.astype(str).str.strip().eq(""), pd.NA)
        return raw

    def _standardize_majority_words(self) -> int:
        fixed = 0
        for col in self.text_cols:
            if col.startswith("__") or col == self.config.notes_column:
                continue
            values = self.df[col].dropna().astype(str).str.strip()
            values = values[~values.str.lower().isin(MISSING_VALUES)]
            if values.empty:
                continue
            mode_vals = values.mode()
            if mode_vals.empty:
                continue
            majority_word = str(mode_vals.iloc[0]).strip()
            if not majority_word:
                continue
            for idx, raw_val in self.df[col].items():
                if pd.isna(raw_val) or not isinstance(raw_val, str):
                    continue
                cell = raw_val.strip()
                if not cell or cell == majority_word or cell.lower() in {"غير محدد", "not specified", "غير محدد / not specified"}:
                    continue
                ratio = difflib.SequenceMatcher(None, cell, majority_word).ratio()
                if ratio >= 0.85:
                    self.df.at[idx, col] = majority_word
                    fixed += 1
        return fixed

    def _build_note_series(self, mask: pd.Series, raw: pd.Series, template: str) -> pd.Series:
        note_series = pd.Series([""] * len(raw), index=raw.index, dtype="string")
        for idx in raw[mask].index:
            note_series.at[idx] = template.format(row=idx, value=raw.at[idx])
        return note_series

    def _append_notes(self, notes: pd.Series) -> None:
        if notes.empty or not notes.any():
            return
        existing = self.df[self.config.notes_column].fillna("").astype("string")
        merged = existing.where(~notes.astype(bool), existing + " | " + notes)
        merged = merged.str.lstrip(" | ")
        self.df[self.config.notes_column] = merged

    def optimize_memory(self) -> Dict[str, Any]:
        for col in self.df.columns:
            if col == self.config.notes_column:
                continue
            series = self.df[col]
            if pd.api.types.is_integer_dtype(series) or pd.api.types.is_float_dtype(series):
                self.df[col] = pd.to_numeric(series, downcast="integer" if pd.api.types.is_integer_dtype(series) else "float")
            elif pd.api.types.is_object_dtype(series):
                cardinality = series.nunique(dropna=False) / max(len(series), 1)
                if cardinality <= self.config.category_threshold:
                    self.df[col] = series.astype("category")

        self.metrics["memory_after_bytes"] = self.df.memory_usage(deep=True).sum()
        reduction = self.metrics["memory_before_bytes"] - self.metrics["memory_after_bytes"]
        logger.info(f"Memory optimized: saved {reduction:,} bytes")
        return {
            "memory_before_bytes": self.metrics["memory_before_bytes"],
            "memory_after_bytes": self.metrics["memory_after_bytes"],
            "reduction_bytes": int(reduction),
        }

    def _normalize_numeric_column(self, raw: pd.Series) -> Tuple[pd.Series, pd.Series]:
        normalized = raw.astype(str).str.strip().replace({"": pd.NA, "nan": pd.NA, "None": pd.NA, "none": pd.NA}, regex=False)
        normalized = normalized.replace({val: pd.NA for val in MISSING_VALUES}, regex=False)
        normalized = normalized.str.replace(r"[,\s]", "", regex=True)
        numeric = pd.to_numeric(normalized, errors="coerce")
        invalid_mask = raw.notna() & numeric.isna()
        return numeric, invalid_mask

    def _normalize_date_column(self, raw: pd.Series) -> Tuple[pd.Series, pd.Series, pd.Series]:
        parsed = pd.to_datetime(raw, errors="coerce", infer_datetime_format=True)
        invalid_mask = raw.notna() & parsed.isna()
        time_mask = parsed.notna() & (parsed.dt.time != datetime.min.time())
        return parsed.dt.normalize(), invalid_mask, time_mask

    def _detect_numeric_outliers(self, numeric: pd.Series, col_name: str) -> List[Tuple[int, str]]:
        coords: List[Tuple[int, str]] = []
        if numeric.notna().sum() < 4:
            return coords
        q1, q3 = numeric.quantile(0.25), numeric.quantile(0.75)
        iqr = q3 - q1
        if iqr == 0:
            return coords
        mask = ((numeric < q1 - 1.5 * iqr) | (numeric > q3 + 1.5 * iqr)) & numeric.notna()
        return [(idx, col_name) for idx in numeric[mask].index]

    def process_numeric_columns(self) -> None:
        logger.info(f"Processing {len(self.config.numeric_columns)} numeric columns")
        for col_name in self.config.numeric_columns:
            if col_name not in self.df.columns:
                logger.warning(f"Numeric column missing: {col_name}")
                continue

            raw = self.df[col_name]
            missing_mask = raw.isna()
            if missing_mask.any():
                self.green_coordinates.extend((idx, col_name) for idx in raw[missing_mask].index)

            numeric, invalid_mask = self._normalize_numeric_column(raw)
            self.yellow_coordinates.extend(self._detect_numeric_outliers(numeric, col_name))

            if invalid_mask.any():
                notes = self._build_note_series(
                    invalid_mask,
                    raw,
                    "[Row{row}] non-numeric value '{value}' removed from " + col_name,
                )
                self._append_notes(notes)
                self.metrics["cells_moved_to_notes"] += int(invalid_mask.sum())
                self.metrics["cells_modified"] += int(invalid_mask.sum())
                self.type_violations.extend(
                    {
                        "row": int(idx),
                        "column": col_name,
                        "expected_type": "numeric",
                        "actual_value": str(raw.at[idx]),
                        "action": "moved_to_notes",
                    }
                    for idx in raw[invalid_mask].index
                )

            integer_mask = numeric.notna() & (numeric % 1 == 0)
            if integer_mask.sum() == numeric.notna().sum() and numeric.notna().any():
                numeric = numeric.astype("Int64")
            else:
                numeric = numeric.astype("Float64")

            if missing_mask.any() and numeric.notna().any():
                non_null = numeric.dropna()
                if len(non_null) > 0:
                    median_val = non_null.median()
                    if not pd.isna(median_val):
                        numeric.loc[missing_mask] = median_val

            self.df[col_name] = numeric
            self.metrics["cells_checked"] += len(raw)
            self.metrics["type_conversions"] += int(numeric.notna().sum())
            self.audit_log.append(
                {
                    "operation": "numeric_processing",
                    "column": col_name,
                    "notes_count": int(invalid_mask.sum()),
                    "timestamp": datetime.now().isoformat(),
                }
            )

    def process_date_columns(self) -> None:
        logger.info(f"Processing {len(self.config.date_columns)} date columns")
        for col_name in self.config.date_columns:
            if col_name not in self.df.columns:
                logger.warning(f"Date column missing: {col_name}")
                continue

            raw = self.df[col_name]
            parsed, invalid_mask, time_mask = self._normalize_date_column(raw)

            if invalid_mask.any():
                notes = self._build_note_series(
                    invalid_mask,
                    raw,
                    "[Row{row}] invalid date '{value}' moved from " + col_name,
                )
                self._append_notes(notes)
                self.metrics["cells_moved_to_notes"] += int(invalid_mask.sum())
                self.metrics["cells_modified"] += int(invalid_mask.sum())
                self.type_violations.extend(
                    {
                        "row": int(idx),
                        "column": col_name,
                        "expected_type": "date",
                        "actual_value": str(raw.at[idx]),
                        "action": "moved_to_notes",
                    }
                    for idx in raw[invalid_mask].index
                )

            if time_mask.any():
                notes = self._build_note_series(
                    time_mask,
                    raw,
                    "[Row{row}] stripped timestamp from '{value}' in " + col_name,
                )
                self._append_notes(notes)
                self.metrics["date_sterilized"] += int(time_mask.sum())
                self.metrics["null_replacements"] += int(time_mask.sum())

            self.df[col_name] = parsed
            self.metrics["cells_checked"] += len(raw)
            self.audit_log.append(
                {
                    "operation": "date_processing",
                    "column": col_name,
                    "notes_count": int(invalid_mask.sum() + time_mask.sum()),
                    "timestamp": datetime.now().isoformat(),
                }
            )

    def process_text_columns(self) -> None:
        logger.info(f"Processing {len(self.text_cols)} text columns")
        for col_name in self.text_cols:
            if col_name not in self.df.columns or col_name == self.config.notes_column:
                continue
            if col_name in self.config.date_columns:
                continue

            raw = self.df[col_name]
            raw = self._drop_blank_strings_for_text(raw)
            missing_mask = raw.isna()
            if missing_mask.any():
                self.red_coordinates.extend((idx, col_name) for idx in raw[missing_mask].index)
                raw.loc[missing_mask] = "غير محدد"

            self.df[col_name] = raw.astype("string")
            self.metrics["cells_checked"] += len(raw)
            self.metrics["cells_modified"] += int(missing_mask.sum())

        fixed = self._standardize_majority_words()
        if fixed > 0:
            logger.info(f"Standardized {fixed} majority word text cells")

    def detect_data_shifting(self) -> None:
        logger.info("Detecting unexpected data shifts")
        compare_columns = [
            c for c in self.original_df.columns
            if c not in self.config.numeric_columns and c not in self.config.date_columns
        ]
        if not compare_columns:
            logger.info("No comparison columns available for shift detection")
            return

        original = self.original_df[compare_columns]
        processed = self.df[compare_columns]
        shift_mask = original.ne(processed) & ~(original.isna() & processed.isna())
        for row_idx, col_idx in zip(*np.where(shift_mask.values)):
            column = shift_mask.columns[col_idx]
            self.data_shift_log.append(
                {
                    "row": int(row_idx),
                    "column": column,
                    "original": str(original.iat[row_idx, col_idx]),
                    "processed": str(processed.iat[row_idx, col_idx]),
                    "timestamp": datetime.now().isoformat(),
                }
            )
        logger.info(f"Detected {len(self.data_shift_log)} unexpected shifts")

    def generate_audit_report(self) -> Dict[str, Any]:
        return {
            "summary": {
                "processing_timestamp": datetime.now().isoformat(),
                "total_rows": self.metrics["total_rows"],
                "total_columns": self.metrics["total_columns"],
                "total_cells": self.metrics["total_cells"],
                "cells_checked": self.metrics["cells_checked"],
                "cells_modified": self.metrics["cells_modified"],
                "cells_moved_to_notes": self.metrics["cells_moved_to_notes"],
                "type_conversions": self.metrics["type_conversions"],
                "date_sterilized": self.metrics["date_sterilized"],
                "null_replacements": self.metrics["null_replacements"],
                "notes_added": int((self.df[self.config.notes_column] != "").sum()),
                "memory_before_bytes": self.metrics["memory_before_bytes"],
                "memory_after_bytes": self.metrics["memory_after_bytes"],
                "coverage_percentage": (
                    self.metrics["cells_checked"] / max(self.metrics["total_cells"], 1)
                ) * 100,
            },
            "operations_log": self.audit_log,
            "type_violations": self.type_violations,
            "data_shifts": self.data_shift_log,
        }

    def export_to_excel(self, filename: str, include_audit_sheet: bool = True) -> None:
        logger.info(f"Exporting to Excel: {filename}")
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            self.df.to_excel(writer, sheet_name="Data", index=False)
            if include_audit_sheet:
                audit_report = self.generate_audit_report()
                pd.DataFrame([audit_report["summary"]]).to_excel(
                    writer, sheet_name=self.config.audit_sheet_name, index=False
                )
                violations_df = pd.DataFrame(audit_report["type_violations"])
                if not violations_df.empty:
                    violations_df.to_excel(writer, sheet_name="Type Violations", index=False)
                shifts_df = pd.DataFrame(audit_report["data_shifts"])
                if not shifts_df.empty:
                    shifts_df.to_excel(writer, sheet_name="Data Shifts", index=False)
        self._apply_excel_formatting(filename)

    def _apply_excel_formatting(self, filename: str) -> None:
        wb = load_workbook(filename)
        ws = wb["Data"]
        numeric_fields = set(self.config.numeric_columns)
        date_fields = set(self.config.date_columns)
        green_coords = set(self.green_coordinates)
        yellow_coords = set(self.yellow_coordinates)
        red_coords = set(self.red_coordinates)

        for col_idx, col_name in enumerate(self.df.columns, start=1):
            if col_name in date_fields:
                for row in range(2, len(self.df) + 2):
                    ws.cell(row=row, column=col_idx).number_format = "YYYY-MM-DD"
            elif col_name in numeric_fields:
                for row in range(2, len(self.df) + 2):
                    value = ws.cell(row=row, column=col_idx).value
                    ws.cell(row=row, column=col_idx).number_format = "0" if isinstance(value, int) else "0.00"
            elif col_name == self.config.notes_column:
                header = ws.cell(row=1, column=col_idx)
                header.font = Font(bold=True)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 50

        header_fill = PatternFill(fill_type="solid", fgColor="D4EDDA")
        outlier_fill = PatternFill(fill_type="solid", fgColor="FFF3CD")
        missing_fill = PatternFill(fill_type="solid", fgColor="F8D7DA")
        for row_idx, row_label in enumerate(self.df.index, start=2):
            for col_idx, col_name in enumerate(self.df.columns, start=1):
                coord = (row_label, col_name)
                cell = ws.cell(row=row_idx, column=col_idx)
                if coord in green_coords:
                    cell.fill = header_fill
                elif coord in yellow_coords:
                    cell.fill = outlier_fill
                elif coord in red_coords:
                    cell.fill = missing_fill
        wb.save(filename)

    def run_complete_pipeline(self, output_file: Optional[str] = None) -> ProcessingResult:
        logger.info("Running complete pipeline")
        self._force_strict_numeric_keywords()
        self._infer_columns_if_missing()
        self._sync_column_sets()
        if self.config.optimize_memory:
            self.optimize_memory()
        start = datetime.now()
        self.process_numeric_columns()
        self.process_date_columns()
        self._sync_column_sets()
        self.process_text_columns()
        self.detect_data_shifting()
        if output_file:
            self.export_to_excel(output_file, include_audit_sheet=self.config.include_audit_sheet)
        duration = (datetime.now() - start).total_seconds()
        self.metrics["memory_after_bytes"] = self.df.memory_usage(deep=True).sum()
        self.metrics["duration_seconds"] = duration
        logger.info(f"Pipeline complete in {duration:.2f}s")
        return ProcessingResult(
            dataframe=self.df,
            audit_report=self.generate_audit_report(),
            duration_seconds=duration,
            status="success",
        )


if __name__ == "__main__":
    sample_data = {
        "ID": ["1", "2.0", "ABC", "4.0", "5"],
        "Name": ["أحمد", "فاطمة", "محمد", "علي", "سارة"],
        "Date": [
            "2024-01-15",
            "2024-02-20 14:30:00",
            "بدون تاريخ",
            "2024-04-10 00:00:00",
            None,
        ],
        "Price": [100, 200.5, 300.0, "غير متوفر", 500.0],
        "Created": [
            "2024-01-01",
            "2024-01-02 08:15:00",
            "2024-01-03",
            "تاريخ غير معروف",
            "2024-01-05",
        ],
    }
    df = pd.DataFrame(sample_data)
    engine = ComprehensiveProcessingEngine(
        df=df,
        config=ProcessingConfig(date_columns=["Date", "Created"], numeric_columns=["ID", "Price"]),
    )
    result = engine.run_complete_pipeline(output_file="processed_sample.xlsx")
    print(json.dumps(result.audit_report, ensure_ascii=False, indent=2))
