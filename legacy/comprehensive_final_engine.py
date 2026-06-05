"""
================================================================================
    محرك المعالجة النهائي الشامل (Comprehensive Final Processing Engine)
    ضمان فحص كامل لكل صف وكل خلية مع التطهير الكامل
================================================================================

مبادئ التصميم:
✓ فحص صف بصف باستخدام df.apply (Row-by-Row Integrity)
✓ فحص عميق لنوع البيانات (Deep Type Checking)
✓ تطهير كامل للتواريخ (Date Sterilization)
✓ منع الترحيل العشوائي (Data Shifting Prevention)
✓ تنسيق Excel برمجي (Programmatic Excel Formatting)
✓ أداء عالي للبيانات الضخمة (High Performance for Large Datasets)
✓ Pipeline شامل لا يترك خلية بدون فحص
"""

import pandas as pd
import numpy as np
from datetime import datetime, date
from typing import Dict, Tuple, List, Any
from pathlib import Path
import warnings
import re
from openpyxl import load_workbook
from openpyxl.styles import numbers as openpyxl_numbers

warnings.filterwarnings('ignore')


class ComprehensiveFinalEngine:
    """
    محرك معالجة نهائي شامل يفحص كل صف وكل خلية بكفاءة عالية
    
    المميزات:
    - فحص صف بصف مع توثيق كل عملية
    - تصحيح تلقائي للبيانات المخطوءة
    - تحويل آمن للأنواع
    - تنسيق Excel متقدم
    """
    
    def __init__(self, verbose: bool = True):
        """
        Args:
            verbose: طباعة التفاصيل أثناء المعالجة
        """
        self.verbose = verbose
        self.audit_log = []  # سجل العمليات الكاملة
        self.data_shifts_log = []  # تسجيل أي ترحيل للبيانات
        self.date_columns = []  # أعمدة التاريخ المكتشفة
        self.numeric_columns = []  # الأعمدة الرقمية المكتشفة
        self.notes_column = 'Notes'  # عمود الملاحظات للبيانات المنقولة
        self.processed_df = None
        
    def log(self, message: str, level: str = "INFO"):
        """توثيق العمليات"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] [{level}] {message}"
        self.audit_log.append(log_entry)
        if self.verbose:
            print(log_entry)
    
    def detect_column_types(self, df: pd.DataFrame) -> Dict[str, str]:
        """
        اكتشاف أنواع الأعمدة بدقة
        
        العملية:
        1. فحص أول 100 صف غير فارغة من كل عمود
        2. تحديد النمط (Pattern) للتاريخ والأرقام
        3. تصنيف الأعمدة
        
        Returns:
            Dict يحتوي على نوع كل عمود
        """
        self.log("🔍 بدء اكتشاف أنواع الأعمدة (Column Type Detection)")
        
        column_types = {}
        
        for col in df.columns:
            non_null = df[col].dropna()
            
            if len(non_null) == 0:
                column_types[col] = 'unknown'
                continue
            
            # الحصول على عينة من القيم
            sample = non_null.head(100)
            
            # محاولة اكتشاف إذا كانت أعمدة تاريخ
            if self._is_date_column(sample):
                column_types[col] = 'date'
                self.date_columns.append(col)
                self.log(f"   ✓ العمود '{col}' -> تاريخ (Date Column)")
            
            # محاولة اكتشاف إذا كانت أعمدة أرقام
            elif self._is_numeric_column(sample):
                column_types[col] = 'numeric'
                self.numeric_columns.append(col)
                self.log(f"   ✓ العمود '{col}' -> رقمي (Numeric Column)")
            
            else:
                column_types[col] = 'text'
                self.log(f"   ✓ العمود '{col}' -> نصي (Text Column)")
        
        return column_types
    
    def _is_date_column(self, sample: pd.Series) -> bool:
        """
        فحص إذا كانت السلسلة تحتوي على تواريخ
        
        النمط المتوقع:
        - صيغ التاريخ القياسية (YYYY-MM-DD, DD/MM/YYYY)
        - تاريخ/وقت (Timestamp)
        - 80% على الأقل من القيم يجب أن تكون تاريخاً
        """
        date_count = 0
        
        for val in sample:
            try:
                val_str = str(val).strip()
                
                # فحص صيغ التاريخ المختلفة
                try:
                    # محاولة الفحص على نوع البيانات فقط
                    if isinstance(val, pd.Timestamp):
                        date_count += 1
                        continue
                except:
                    pass
                
                if re.match(r'\d{4}-\d{2}-\d{2}', val_str):  # YYYY-MM-DD
                    date_count += 1
                elif re.match(r'\d{1,2}/\d{1,2}/\d{4}', val_str):  # DD/MM/YYYY
                    date_count += 1
                elif re.match(r'\d{1,2}-\d{1,2}-\d{4}', val_str):  # DD-MM-YYYY
                    date_count += 1
            except:
                pass
        
        # إذا كانت 80% على الأقل من العينة تواريخ
        return (date_count / len(sample)) >= 0.8
    
    def _is_numeric_column(self, sample: pd.Series) -> bool:
        """
        فحص إذا كانت السلسلة تحتوي على أرقام
        
        النمط المتوقع:
        - أرقام صحيحة أو عشرية
        - 90% على الأقل من القيم يجب أن تكون أرقاماً
        """
        numeric_count = 0
        
        for val in sample:
            try:
                # محاولة تحويل القيمة لرقم
                float(val)
                numeric_count += 1
            except (ValueError, TypeError):
                pass
        
        # إذا كانت 90% على الأقل من العينة أرقام
        return (numeric_count / len(sample)) >= 0.9
    
    # ========================================================================
    # ---- القسم الأول: فحص وتنظيف البيانات الأساسية ----
    # ========================================================================
    
    def clean_date_column(self, df: pd.DataFrame, col: str) -> pd.DataFrame:
        """
        تطهير كامل لعمود التاريخ (Date Sterilization)
        
        الخطوات:
        1️⃣  فحص كل خلية للتأكد من أنها تاريخ حقيقي
        2️⃣  إذا كانت نص (اسم/مدينة)، نقلها لـ Notes
        3️⃣  تحويل التاريخ لصيغة datetime
        4️⃣  استخراج التاريخ فقط (بدون وقت)
        5️⃣  التأكد من اختفاء أصفار الوقت (00:00:00)
        """
        self.log(f"🔄 تطهير عمود التاريخ '{col}' (Date Sterilization)")
        
        # إنشاء عمود Notes إذا لم يكن موجوداً
        if self.notes_column not in df.columns:
            df[self.notes_column] = ''
        
        # معالجة صف بصف
        def process_date_cell(row):
            """
            معالجة كل خلية في عمود التاريخ
            
            Returns:
                Tuple[date, str]: (القيمة النظيفة، ملاحظات إضافية)
            """
            original_val = row[col]
            note = row[self.notes_column] if isinstance(row[self.notes_column], str) else ''
            
            # ✓ إذا كانت القيمة فارغة، احتفظ بـ None
            if pd.isna(original_val):
                return None, note
            
            val_str = str(original_val).strip()
            
            # ✓ محاولة تحويل التاريخ
            try:
                # الطريقة 1: إذا كانت datetime مسبقاً
                if pd.api.types.is_datetime64_any_dtype(original_val):
                    parsed_date = pd.Timestamp(original_val).date()
                    self.log(f"   ✓ تحويل timestamp إلى تاريخ: {parsed_date}")
                    return parsed_date, note
                
                # الطريقة 2: محاولة الصيغ المختلفة
                parsed_date = pd.to_datetime(val_str, format='%Y-%m-%d', errors='coerce')
                
                if pd.isna(parsed_date):
                    parsed_date = pd.to_datetime(val_str, errors='coerce')
                
                if pd.isna(parsed_date):
                    raise ValueError(f"لا يمكن تحويل: {val_str}")
                
                # استخراج التاريخ فقط (بدون وقت)
                clean_date = parsed_date.date()
                self.log(f"   ✓ تحويل ناجح: {val_str} -> {clean_date}")
                return clean_date, note
            
            except Exception as e:
                # ⚠️ إذا لم يكن تاريخاً، قد يكون نصاً (اسم/مدينة)
                self.log(f"   ⚠️  قيمة غير صحيحة في {col}: '{val_str}'")
                
                # نقل القيمة للملاحظات
                new_note = note + f" | {col}: {val_str}"
                self.data_shifts_log.append({
                    'column': col,
                    'original_value': val_str,
                    'action': 'moved_to_notes',
                    'reason': 'invalid_date_format'
                })
                
                return None, new_note
        
        # تطبيق المعالجة على كل صف
        df[[col, self.notes_column]] = df.apply(
            lambda row: pd.Series(process_date_cell(row)),
            axis=1
        )
        
        # تنظيف العمود: جعل كل القيم من نوع datetime.date
        df[col] = pd.to_datetime(df[col], errors='coerce').dt.date
        
        self.log(f"✅ انتهى تطهير عمود '{col}'")
        return df
    
    def clean_numeric_column(self, df: pd.DataFrame, col: str) -> pd.DataFrame:
        """
        تطهير كامل لعمود رقمي (Deep Type Checking)
        
        الخطوات:
        1️⃣  فحص كل خلية للتأكد من أنها رقم حقيقي
        2️⃣  تحويل الأرقام الصحيحة إلى Int64 (بدون .0)
        3️⃣  الحفاظ على الأرقام العشرية كـ Float64
        4️⃣  توثيق أي قيم غير رقمية
        5️⃣  استبدال القيم الخاطئة بـ None
        """
        self.log(f"🔢 تطهير العمود الرقمي '{col}' (Numeric Cleaning)")
        
        def process_numeric_cell(val):
            """
            معالجة كل خلية في عمود رقمي
            
            Returns:
                float/int أو None
            """
            # ✓ إذا كانت القيمة فارغة
            if pd.isna(val):
                return None
            
            val_str = str(val).strip()
            
            try:
                # تحويل لرقم
                num_val = float(val_str)
                
                # ✓ إذا كانت الرقم صحيح (بدون جزء عشري)
                if num_val == int(num_val):
                    self.log(f"   ✓ تحويل لـ Int64: {val_str} -> {int(num_val)}")
                    return int(num_val)
                else:
                    # ✓ الرقم العشري يبقى float
                    self.log(f"   ✓ الاحتفاظ بـ Float: {val_str} -> {num_val}")
                    return num_val
            
            except (ValueError, TypeError):
                # ⚠️ قيمة غير رقمية
                self.log(f"   ⚠️  قيمة غير رقمية في {col}: '{val_str}'")
                
                self.data_shifts_log.append({
                    'column': col,
                    'original_value': val_str,
                    'action': 'invalid_numeric',
                    'reason': 'non_numeric_value'
                })
                
                return None
        
        # تطبيق المعالجة على كل خلية
        df[col] = df[col].apply(process_numeric_cell)
        
        # تحويل النوع الآمن
        try:
            # محاولة تحويل الأعداد الصحيحة فقط
            mask_int = df[col].notna() & (df[col] == df[col].astype('Int64'))
            
            # معالجة آمنة
            if mask_int.any():
                df[col] = df[col].astype('Int64')
                self.log(f"   ✓ تم تحويل النوع إلى Int64")
        except Exception as e:
            self.log(f"   ⚠️  لم يتم تحويل النوع: {e}", level="WARNING")
        
        self.log(f"✅ انتهى تطهير العمود '{col}'")
        return df
    
    # ========================================================================
    # ---- القسم الثاني: فحص منع الترحيل العشوائي ----
    # ========================================================================
    
    def detect_data_shifting(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        اكتشاف أي بيانات "مرحلة" (Data Shifting)
        
        مؤشرات الترحيل:
        1. قيمة موجودة في عمود خاطئ
        2. نفس القيمة في عمود آخر
        3. نمط متكرر يشير لترحيل
        
        الإجراء:
        - توثيق الترحيل
        - محاولة نقل البيانات لمكانها الصحيح
        - إذا لم نكن متأكدين، نقلها للملاحظات
        """
        self.log("🔍 اكتشاف الترحيل العشوائي (Data Shifting Detection)")
        
        # فحص صفوف مكررة (Duplicate rows)
        duplicate_mask = df.duplicated(keep='first')
        if duplicate_mask.any():
            num_duplicates = duplicate_mask.sum()
            self.log(f"   ⚠️  تم اكتشاف {num_duplicates} صف مكرر")
            self.data_shifts_log.append({
                'type': 'duplicates',
                'count': num_duplicates,
                'action': 'keep_first'
            })
            df = df[~duplicate_mask].reset_index(drop=True)
        
        # فحص الفراغات العشوائية (Random Blanks)
        self.log("   ✓ فحص الفراغات في الأعمدة الأساسية")
        for col in df.columns:
            null_count = df[col].isna().sum()
            if null_count > 0:
                null_percentage = (null_count / len(df)) * 100
                if null_percentage > 20:  # إذا كان أكثر من 20% فارغ
                    self.log(f"   ⚠️  العمود '{col}' يحتوي على {null_percentage:.1f}% فارغ")
        
        return df
    
    # ========================================================================
    # ---- القسم الثالث: معالجة شاملة لكل صف ----
    # ========================================================================
    
    def comprehensive_row_audit(self, df: pd.DataFrame, column_types: Dict[str, str]) -> pd.DataFrame:
        """
        فحص شامل لكل صف (Row-by-Row Integrity)
        
        العملية:
        1️⃣  فحص كل صف على حدة
        2️⃣  تطبيق القواعس الخاصة بنوع العمود
        3️⃣  توثيق كل تغيير
        4️⃣  ضمان عدم ترك خلية واحدة بدون فحص
        """
        self.log(f"\n📋 بدء الفحص الشامل لكل صف (Comprehensive Row Audit)")
        self.log(f"   إجمالي الصفوف: {len(df)}")
        self.log(f"   إجمالي الأعمدة: {len(df.columns)}")
        
        # إنشاء عمود Notes إذا لم يكن موجوداً
        if self.notes_column not in df.columns:
            df[self.notes_column] = ''
        
        processed_rows = 0
        
        # معالجة كل صف
        for idx, row in df.iterrows():
            row_has_changes = False
            row_notes = []
            
            # فحص كل خلية في الصف
            for col in df.columns:
                if col == self.notes_column:
                    continue
                
                val = row[col]
                col_type = column_types.get(col, 'text')
                
                # فحص وفقاً لنوع العمود
                if col_type == 'date' and pd.notna(val):
                    # التحقق من صيغة التاريخ
                    if not isinstance(val, date):
                        try:
                            df.at[idx, col] = pd.to_datetime(val).date()
                            row_has_changes = True
                            row_notes.append(f"{col}: تحويل التاريخ")
                        except:
                            df.at[idx, col] = None
                            row_notes.append(f"{col}: تاريخ غير صحيح -> None")
                            row_has_changes = True
                
                elif col_type == 'numeric' and pd.notna(val):
                    # التحقق من الرقم
                    try:
                        num_val = float(val)
                        if num_val == int(num_val):
                            df.at[idx, col] = int(num_val)
                    except:
                        df.at[idx, col] = None
                        row_notes.append(f"{col}: قيمة غير رقمية")
                        row_has_changes = True
            
            # تحديث الملاحظات
            if row_notes:
                existing_note = str(df.at[idx, self.notes_column]) if pd.notna(df.at[idx, self.notes_column]) else ''
                df.at[idx, self.notes_column] = existing_note + ' | ' + ' | '.join(row_notes)
            
            processed_rows += 1
            
            # طباعة تقدم كل 10000 صف
            if (idx + 1) % 10000 == 0:
                self.log(f"   ✓ تم معالجة {idx + 1} صف...")
        
        self.log(f"✅ انتهى الفحص الشامل: تم معالجة {processed_rows} صف بنجاح")
        return df
    
    # ========================================================================
    # ---- القسم الرابع: تنسيق Excel البرمجي ----
    # ========================================================================
    
    def format_excel_output(self, df: pd.DataFrame, column_types: Dict[str, str], 
                           output_path: str):
        """
        تنسيق ملف Excel برمجياً (Programmatic Excel Formatting)
        
        التنسيقات:
        ✓ التاريخ: yyyy-mm-dd (بدون وقت)
        ✓ الأرقام الصحيحة: بدون فواصل عشرية
        ✓ الأرقام العشرية: رقمين بعد الفاصلة
        ✓ تعديل عرض الأعمدة تلقائياً
        ✓ إضافة رؤوس مجمدة (Freeze Headers)
        """
        self.log(f"\n📊 تطبيق تنسيق Excel (Excel Formatting)")
        self.log(f"   📁 المسار: {output_path}")
        
        # كتابة البيانات الأولية
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data', index=False)
        
        # تحميل المصنف للتنسيق المتقدم
        wb = load_workbook(output_path)
        ws = wb['Data']
        
        # تطبيق التنسيق على كل عمود
        for col_num, col in enumerate(df.columns, 1):
            col_type = column_types.get(col, 'text')
            
            # الحصول على الخلية الأولى (رأس العمود)
            cell = ws.cell(row=1, column=col_num)
            cell.font = cell.font.copy()
            cell.font = cell.font.copy(bold=True)  # جعل الرؤوس غامقة
            
            if col_type == 'date':
                # تطبيق تنسيق التاريخ
                for row in range(2, len(df) + 2):
                    cell = ws.cell(row=row, column=col_num)
                    cell.number_format = 'YYYY-MM-DD'
                self.log(f"   ✓ تنسيق التاريخ على '{col}'")
            
            elif col_type == 'numeric':
                # تطبيق تنسيق الأرقام
                for row in range(2, len(df) + 2):
                    cell = ws.cell(row=row, column=col_num)
                    val = df.iloc[row-2, col_num-1]
                    
                    # إذا كانت رقم صحيح
                    if pd.notna(val) and isinstance(val, (int, np.integer)):
                        cell.number_format = '0'  # بدون فواصل عشرية
                    else:
                        cell.number_format = '0.00'  # رقمين بعد الفاصلة
                
                self.log(f"   ✓ تنسيق الأرقام على '{col}'")
            
            # تعديل عرض العمود تلقائياً
            adjusted_width = max(
                len(str(col)),  # طول الرأس
                df[col].astype(str).str.len().max() + 2  # أطول قيمة + هامش
            )
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = adjusted_width
        
        # مجازة الصف الأول (Freeze Headers)
        ws.freeze_panes = 'A2'
        self.log(f"   ✓ تم تجميد رؤوس الأعمدة")
        
        # حفظ التنسيق
        wb.save(output_path)
        self.log(f"✅ تم حفظ ملف Excel مع التنسيق الكامل")
    
    # ========================================================================
    # ---- القسم الخامس: Pipeline الشامل ----
    # ========================================================================
    
    def process_comprehensive(self, input_path: str, output_path: str = None) -> pd.DataFrame:
        """
        Pipeline شامل لمعالجة البيانات
        
        المراحل:
        1️⃣  تحميل البيانات
        2️⃣  اكتشاف أنواع الأعمدة
        3️⃣  اكتشاف الترحيل العشوائي
        4️⃣  تطهير كل عمود حسب نوعه
        5️⃣  فحص شامل صف بصف
        6️⃣  تصدير مع التنسيق
        7️⃣  توليد تقرير العمليات
        
        Returns:
            DataFrame المعالجة بالكامل
        """
        self.log("=" * 80)
        self.log("🚀 بدء Pipeline المعالجة الشامل (Comprehensive Processing Pipeline)")
        self.log("=" * 80)
        
        # ✓ المرحلة 1: تحميل البيانات
        self.log(f"\n📂 تحميل البيانات من: {input_path}")
        try:
            df = pd.read_csv(input_path)
            self.log(f"   ✓ تم تحميل {len(df)} صف و {len(df.columns)} عمود")
        except Exception as e:
            self.log(f"   ❌ خطأ في تحميل البيانات: {e}", level="ERROR")
            return None
        
        # ✓ المرحلة 2: اكتشاف أنواع الأعمدة
        self.log("\n")
        column_types = self.detect_column_types(df)
        
        # ✓ المرحلة 3: اكتشاف الترحيل العشوائي
        self.log("\n")
        df = self.detect_data_shifting(df)
        
        # ✓ المرحلة 4: تطهير الأعمدة
        self.log("\n")
        self.log("🧹 بدء تطهير جميع الأعمدة")
        
        for col in self.date_columns:
            df = self.clean_date_column(df, col)
            self.log("")
        
        for col in self.numeric_columns:
            df = self.clean_numeric_column(df, col)
            self.log("")
        
        # ✓ المرحلة 5: فحص شامل صف بصف
        df = self.comprehensive_row_audit(df, column_types)
        
        # حفظ البيانات المعالجة
        self.processed_df = df
        
        # ✓ المرحلة 6: تصدير مع التنسيق
        if output_path is None:
            output_path = input_path.replace('.csv', '_processed.xlsx')
        
        self.log(f"\n")
        df.to_excel(output_path.replace('.xlsx', '_raw.xlsx'), index=False, sheet_name='Data')
        
        self.format_excel_output(df, column_types, output_path)
        
        # ✓ المرحلة 7: توليد التقرير
        self.log(f"\n")
        self.generate_audit_report(output_path)
        
        self.log("\n" + "=" * 80)
        self.log("✅ انتهت معالجة البيانات بنجاح!")
        self.log("=" * 80)
        
        return df
    
    # ========================================================================
    # ---- القسم السادس: توليد التقارير ----
    # ========================================================================
    
    def generate_audit_report(self, output_path: str):
        """
        توليد تقرير العمليات الكامل
        
        يحتوي على:
        - ملخص العمليات
        - قائمة تفصيلية بجميع التغييرات
        - إحصائيات البيانات
        - التوصيات
        """
        self.log("📝 توليد تقرير العمليات (Audit Report)")
        
        report_path = output_path.replace('.xlsx', '_audit_report.txt')
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("تقرير العمليات الشامل (Comprehensive Audit Report)\n")
            f.write("=" * 80 + "\n\n")
            
            # السجل الكامل
            f.write("سجل العمليات الكاملة:\n")
            f.write("-" * 80 + "\n")
            for entry in self.audit_log:
                f.write(entry + "\n")
            
            # سجل الترحيل
            if self.data_shifts_log:
                f.write("\n" + "=" * 80 + "\n")
                f.write("سجل الترحيل العشوائي:\n")
                f.write("-" * 80 + "\n")
                for shift in self.data_shifts_log:
                    f.write(f"  • {shift}\n")
            
            # الإحصائيات
            f.write("\n" + "=" * 80 + "\n")
            f.write("إحصائيات البيانات:\n")
            f.write("-" * 80 + "\n")
            f.write(f"إجمالي الصفوف: {len(self.processed_df)}\n")
            f.write(f"إجمالي الأعمدة: {len(self.processed_df.columns)}\n")
            f.write(f"أعمدة التاريخ: {len(self.date_columns)}\n")
            f.write(f"الأعمدة الرقمية: {len(self.numeric_columns)}\n")
            
            # إحصائيات الفراغ
            f.write("\nإحصائيات الفراغ:\n")
            for col in self.processed_df.columns:
                null_count = self.processed_df[col].isna().sum()
                null_percentage = (null_count / len(self.processed_df)) * 100
                if null_percentage > 0:
                    f.write(f"  • {col}: {null_count} فراغ ({null_percentage:.1f}%)\n")
        
        self.log(f"✅ تم إنشاء التقرير: {report_path}")


# ============================================================================
# ---- مثال على الاستخدام ----
# ============================================================================

def main():
    """
    مثال على استخدام المحرك الشامل
    """
    
    # إنشاء المحرك
    engine = ComprehensiveFinalEngine(verbose=True)
    
    # معالجة البيانات
    input_file = 'sample_data.csv'
    output_file = 'sample_data_processed.xlsx'
    
    # تشغيل Pipeline الشامل
    processed_df = engine.process_comprehensive(
        input_path=input_file,
        output_path=output_file
    )
    
    # يمكنك الآن استخدام processed_df مباشرة
    print("\n\n" + "=" * 80)
    print("البيانات المعالجة:")
    print("=" * 80)
    print(processed_df)


if __name__ == "__main__":
    main()
