"""
محرك المعالجة الشاملة (Comprehensive Processing Engine)
هذا الملف يوفر خط إنتاج متكامل للفحص الدقيق لكل صف وخلية في DataFrames
مع ضمان عدم تجاهل أي بيانات عشوائية وتطبيق معايير جودة عالية جداً

Version: 1.0
Author: Data Quality Team
"""

import pandas as pd
import numpy as np
from datetime import datetime, date
from typing import Dict, List, Tuple, Any, Optional
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import json

# إعداد نظام التسجيل (Logging)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ComprehensiveProcessingEngine:
    """
    محرك معالجة شامل يضمن فحص كل صف وخلية في DataFrame
    
    المبادئ الأساسية:
    1. Row-by-Row Integrity: فحص كل صف بشكل مستقل
    2. Deep Type Checking: التحقق من أنواع البيانات
    3. Date Sterilization: تنظيف التواريخ من أصفار الوقت
    4. Data Shifting Prevention: توثيق أي ترحيل بيانات
    5. High Performance: معالجة سريعة حتى مع ملايين الصفوف
    """
    
    def __init__(self, df: pd.DataFrame, date_columns: List[str] = None, 
                 numeric_columns: List[str] = None):
        """
        تهيئة محرك المعالجة
        
        Parameters:
        -----------
        df : pd.DataFrame
            البيانات المراد معالجتها
        date_columns : List[str]
            قائمة أسماء أعمدة التاريخ
        numeric_columns : List[str]
            قائمة أسماء الأعمدة الرقمية
        """
        self.original_df = df.copy()  # الحفاظ على البيانات الأصلية للمقارنة
        self.df = df.copy()  # نسخة العمل
        self.date_columns = date_columns or []
        self.numeric_columns = numeric_columns or []
        
        # إنشاء عمود Notes لتسجيل جميع التغييرات
        if 'Notes' not in self.df.columns:
            self.df['Notes'] = ''
        
        # سجل شامل لجميع العمليات (Audit Log)
        self.audit_log = []
        self.data_shift_log = []
        self.type_violations = []
        self.stats = {
            'total_rows': len(df),
            'total_cells': len(df) * len(df.columns),
            'cells_checked': 0,
            'cells_modified': 0,
            'cells_moved_to_notes': 0,
            'type_conversions': 0,
            'date_sterilized': 0,
            'null_replacements': 0
        }
        
        logger.info(f"تم تهيئة محرك المعالجة: {len(df)} صف × {len(df.columns)} عمود")
    
    # ============================================================
    # المرحلة 1: فحص نوع البيانات العميق (Deep Type Checking)
    # ============================================================
    
    def _is_valid_date(self, value) -> bool:
        """فحص ما إذا كانت القيمة تاريخاً صحيحاً"""
        if pd.isna(value):
            return False
        if isinstance(value, (datetime, date)):
            return True
        if isinstance(value, str):
            try:
                pd.to_datetime(value, errors='coerce')
                parsed = pd.to_datetime(value, errors='coerce')
                return pd.notna(parsed)
            except:
                return False
        return False
    
    def _is_numeric_integer(self, value) -> bool:
        """فحص ما إذا كانت القيمة رقماً صحيحاً بدون أجزاء عشرية"""
        if pd.isna(value):
            return False
        try:
            num = float(value)
            # التحقق من أن الرقم صحيح (بدون جزء عشري)
            return num == int(num)
        except (ValueError, TypeError):
            return False
    
    def _is_text_in_numeric_column(self, value) -> bool:
        """فحص ما إذا كانت خلية في عمود رقمي تحتوي على نص"""
        if pd.isna(value):
            return False
        if isinstance(value, str):
            # فحص إذا كانت النصوص الشائعة (أسماء، مدن، إلخ)
            try:
                float(value)
                return False  # إذا كانت قابلة للتحويل لرقم، فهي ليست نص
            except ValueError:
                return True  # نص حقيقي
        return False
    
    def _is_text_in_date_column(self, value) -> bool:
        """فحص ما إذا كانت خلية في عمود تاريخ تحتوي على نص عادي"""
        if pd.isna(value):
            return False
        if isinstance(value, str):
            # محاولة تحويلها لتاريخ
            try:
                pd.to_datetime(value, errors='coerce')
                parsed = pd.to_datetime(value, errors='coerce')
                if pd.isna(parsed):
                    return True  # نص حقيقي، ليس تاريخاً
                return False
            except:
                return True
        return False
    
    # ============================================================
    # المرحلة 2: معالجة أعمدة الأرقام (Numeric Column Processing)
    # ============================================================
    
    def _process_numeric_cell(self, value: Any, row_idx: int, col_name: str) -> Tuple[Any, str]:
        """
        معالجة خلية رقمية واحدة مع ضمان فحص دقيق
        
        Returns:
        --------
        (processed_value, note): القيمة المعالجة والملاحظة
        """
        note = ""
        
        # الحالة 1: القيمة فارغة - لا تغيير
        if pd.isna(value):
            self.stats['cells_checked'] += 1
            return value, ""
        
        # الحالة 2: نص في عمود رقمي - نقل للملاحظات
        if self._is_text_in_numeric_column(value):
            self.stats['cells_checked'] += 1
            self.stats['cells_moved_to_notes'] += 1
            note = f"[Row{row_idx}] محتوى نصي '{value}' تم نقله من {col_name}"
            self.type_violations.append({
                'row': row_idx,
                'column': col_name,
                'expected_type': 'numeric',
                'actual_value': value,
                'action': 'moved_to_notes'
            })
            return np.nan, note
        
        # الحالة 3: رقم صحيح - تحويل لـ Int64 (بدون .0)
        try:
            if self._is_numeric_integer(value):
                self.stats['cells_checked'] += 1
                self.stats['type_conversions'] += 1
                int_value = int(float(value))
                return int_value, ""
            else:
                # رقم عشري - الحفاظ عليه
                self.stats['cells_checked'] += 1
                return float(value), ""
        except (ValueError, TypeError):
            self.stats['cells_checked'] += 1
            note = f"[Row{row_idx}] فشل تحويل '{value}' في {col_name}"
            return np.nan, note
    
    def process_numeric_columns(self) -> None:
        """
        معالجة جميع الأعمدة الرقمية صف بصف
        ✓ فحص كل خلية بشكل مستقل
        ✓ نقل النصوص للملاحظات
        ✓ تحويل الأرقام الصحيحة لـ Int64
        """
        logger.info(f"بدء معالجة {len(self.numeric_columns)} أعمدة رقمية...")
        
        for col_name in self.numeric_columns:
            if col_name not in self.df.columns:
                logger.warning(f"العمود '{col_name}' غير موجود")
                continue
            
            processed_values = []
            notes_list = []
            
            # معالجة كل صف بشكل مستقل (Row-by-Row Processing)
            for idx, value in enumerate(self.df[col_name]):
                processed_value, note = self._process_numeric_cell(value, idx, col_name)
                processed_values.append(processed_value)
                
                if note:
                    notes_list.append(note)
                    # إضافة الملاحظة للعمود Notes
                    if pd.isna(self.df.loc[idx, 'Notes']) or self.df.loc[idx, 'Notes'] == '':
                        self.df.loc[idx, 'Notes'] = note
                    else:
                        self.df.loc[idx, 'Notes'] += f" | {note}"
                    
                    self.stats['cells_modified'] += 1
            
            # تحديث العمود برمته
            self.df[col_name] = processed_values
            
            # تحويل لـ Int64 إذا كانت جميع القيم أرقام صحيحة أو فارغة
            try:
                if all(pd.isna(v) or isinstance(v, (int, np.integer)) for v in processed_values):
                    self.df[col_name] = self.df[col_name].astype('Int64')  # Int64 يحافظ على NaN
            except:
                pass
            
            logger.info(f"✓ تم معالجة العمود '{col_name}': {len(notes_list)} ملاحظة")
            self.audit_log.append({
                'operation': 'numeric_processing',
                'column': col_name,
                'notes_count': len(notes_list),
                'timestamp': datetime.now().isoformat()
            })
    
    # ============================================================
    # المرحلة 3: معالجة أعمدة التاريخ (Date Column Processing)
    # ============================================================
    
    def _process_date_cell(self, value: Any, row_idx: int, col_name: str) -> Tuple[Any, str]:
        """
        معالجة خلية تاريخ واحدة مع ضمان فحص دقيق
        
        Returns:
        --------
        (processed_value, note): القيمة المعالجة والملاحظة
        """
        note = ""
        
        # الحالة 1: القيمة فارغة
        if pd.isna(value):
            self.stats['cells_checked'] += 1
            return None, ""
        
        # الحالة 2: نص في عمود تاريخ - نقل للملاحظات
        if self._is_text_in_date_column(value):
            self.stats['cells_checked'] += 1
            self.stats['cells_moved_to_notes'] += 1
            note = f"[Row{row_idx}] نص '{value}' تم نقله من {col_name}"
            self.type_violations.append({
                'row': row_idx,
                'column': col_name,
                'expected_type': 'date',
                'actual_value': value,
                'action': 'moved_to_notes'
            })
            return None, note
        
        # الحالة 3: تاريخ صحيح - تنظيف وتحويل
        try:
            parsed_date = pd.to_datetime(value, errors='coerce')
            
            if pd.notna(parsed_date):
                self.stats['cells_checked'] += 1
                self.stats['date_sterilized'] += 1
                
                # استخراج الجزء الخاص بالتاريخ فقط (بدون الوقت)
                clean_date = parsed_date.date()
                
                # فحص إذا كان هناك وقت غير صفري
                if parsed_date.time() != datetime.min.time():
                    note = f"[Row{row_idx}] تم إزالة الوقت: {parsed_date.time()} من {col_name}"
                    self.stats['null_replacements'] += 1
                
                return clean_date, note
            else:
                self.stats['cells_checked'] += 1
                note = f"[Row{row_idx}] فشل تحويل '{value}' في {col_name}"
                return None, note
                
        except Exception as e:
            self.stats['cells_checked'] += 1
            note = f"[Row{row_idx}] خطأ: {str(e)} في {col_name}"
            return None, note
    
    def process_date_columns(self) -> None:
        """
        معالجة جميع أعمدة التاريخ صف بصف
        ✓ فحص كل خلية بشكل مستقل
        ✓ نقل النصوص للملاحظات
        ✓ تنظيف أصفار الوقت (00:00:00)
        """
        logger.info(f"بدء معالجة {len(self.date_columns)} أعمدة تاريخ...")
        
        for col_name in self.date_columns:
            if col_name not in self.df.columns:
                logger.warning(f"العمود '{col_name}' غير موجود")
                continue
            
            processed_values = []
            notes_list = []
            
            # معالجة كل صف بشكل مستقل
            for idx, value in enumerate(self.df[col_name]):
                processed_value, note = self._process_date_cell(value, idx, col_name)
                processed_values.append(processed_value)
                
                if note:
                    notes_list.append(note)
                    # إضافة الملاحظة للعمود Notes
                    if pd.isna(self.df.loc[idx, 'Notes']) or self.df.loc[idx, 'Notes'] == '':
                        self.df.loc[idx, 'Notes'] = note
                    else:
                        self.df.loc[idx, 'Notes'] += f" | {note}"
                    
                    self.stats['cells_modified'] += 1
            
            # تحديث العمود برمته
            self.df[col_name] = processed_values
            
            logger.info(f"✓ تم معالجة العمود '{col_name}': {len(notes_list)} ملاحظة")
            self.audit_log.append({
                'operation': 'date_processing',
                'column': col_name,
                'notes_count': len(notes_list),
                'timestamp': datetime.now().isoformat()
            })
    
    # ============================================================
    # المرحلة 4: فحص الترحيل العشوائي (Data Shifting Detection)
    # ============================================================
    
    def detect_data_shifting(self) -> None:
        """
        فحص شامل للبيانات المرحلة بين الأعمدة
        ✓ مقارنة البيانات الأصلية بالمعالجة
        ✓ توثيق أي تغييرات غير متوقعة
        """
        logger.info("بدء فحص الترحيل العشوائي...")
        
        for col in self.original_df.columns:
            if col not in self.df.columns:
                continue
            
            # مقارنة القيم
            for idx in range(len(self.original_df)):
                original_value = self.original_df.iloc[idx][col]
                processed_value = self.df.iloc[idx][col]
                
                # إذا كانت القيم مختلفة وليست null
                if not (pd.isna(original_value) and pd.isna(processed_value)):
                    if original_value != processed_value:
                        # فحص إذا كان التغيير متوقعاً
                        if col not in self.numeric_columns and col not in self.date_columns:
                            self.data_shift_log.append({
                                'row': idx,
                                'column': col,
                                'original': str(original_value),
                                'processed': str(processed_value),
                                'timestamp': datetime.now().isoformat()
                            })
        
        logger.info(f"✓ تم فحص الترحيل: {len(self.data_shift_log)} تغييراً غير متوقع")
    
    # ============================================================
    # المرحلة 5: إنشاء سجل العمليات الشامل (Comprehensive Audit Log)
    # ============================================================
    
    def generate_audit_report(self) -> Dict:
        """
        إنشاء تقرير شامل لجميع العمليات
        """
        report = {
            'summary': {
                'processing_timestamp': datetime.now().isoformat(),
                'total_rows': self.stats['total_rows'],
                'total_cells': self.stats['total_cells'],
                'cells_checked': self.stats['cells_checked'],
                'cells_modified': self.stats['cells_modified'],
                'cells_moved_to_notes': self.stats['cells_moved_to_notes'],
                'type_conversions': self.stats['type_conversions'],
                'date_sterilized': self.stats['date_sterilized'],
                'null_replacements': self.stats['null_replacements'],
                'coverage_percentage': (self.stats['cells_checked'] / max(self.stats['total_cells'], 1)) * 100
            },
            'operations_log': self.audit_log,
            'type_violations': self.type_violations,
            'data_shifts': self.data_shift_log,
            'notes_added': (self.df['Notes'] != '').sum()
        }
        
        return report
    
    # ============================================================
    # المرحلة 6: تصدير لـ Excel مع التنسيق الصحيح
    # ============================================================
    
    def export_to_excel(self, filename: str, include_audit_sheet: bool = True) -> None:
        """
        تصدير البيانات المعالجة لـ Excel مع التنسيق الدقيق
        
        ✓ تنسيق أعمدة التاريخ بشكل صحيح
        ✓ تنسيق الأعمدة الرقمية بدون علامات زائدة
        ✓ إضافة ورقة سجل العمليات
        """
        logger.info(f"بدء التصدير لـ Excel: {filename}")
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # كتابة البيانات الرئيسية
            self.df.to_excel(writer, sheet_name='Data', index=False)
            
            # إضافة ورقة سجل العمليات إذا تم الطلب
            if include_audit_sheet:
                audit_report = self.generate_audit_report()
                
                # إنشاء DataFrame من التقرير
                audit_df = pd.DataFrame([audit_report['summary']])
                audit_df.to_excel(writer, sheet_name='Audit Report', index=False)
                
                # كتابة التفاصيل
                violations_df = pd.DataFrame(audit_report['type_violations'])
                if len(violations_df) > 0:
                    violations_df.to_excel(writer, sheet_name='Type Violations', index=False)
                
                data_shifts_df = pd.DataFrame(audit_report['data_shifts'])
                if len(data_shifts_df) > 0:
                    data_shifts_df.to_excel(writer, sheet_name='Data Shifts', index=False)
        
        # تطبيق التنسيق الدقيق
        self._apply_excel_formatting(filename)
        logger.info(f"✓ تم التصدير بنجاح إلى: {filename}")
    
    def _apply_excel_formatting(self, filename: str) -> None:
        """
        تطبيق التنسيق الدقيق على ملف Excel
        """
        wb = load_workbook(filename)
        ws = wb['Data']
        
        # تطبيق التنسيق على كل عمود
        for col_idx, col_name in enumerate(self.df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            
            # تنسيق أعمدة التاريخ
            if col_name in self.date_columns:
                # تطبيق تنسيق التاريخ
                for row in range(2, len(self.df) + 2):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.number_format = 'YYYY-MM-DD'
            
            # تنسيق الأعمدة الرقمية
            elif col_name in self.numeric_columns:
                for row in range(2, len(self.df) + 2):
                    cell = ws.cell(row=row, column=col_idx)
                    # فحص إذا كانت قيمة صحيحة
                    if isinstance(cell.value, (int, np.integer)):
                        cell.number_format = '0'  # بدون فاصلة عشرية
                    else:
                        cell.number_format = '0.00'
            
            # تنسيق عمود الملاحظات
            elif col_name == 'Notes':
                cell.font = Font(italic=True, color='FF0000')
                ws.column_dimensions[cell.column_letter].width = 50
        
        wb.save(filename)
    
    # ============================================================
    # تشغيل الخط الكامل (Run Complete Pipeline)
    # ============================================================
    
    def run_complete_pipeline(self, output_file: str = None) -> Dict:
        """
        تشغيل خط الإنتاج الكامل من البداية للنهاية
        
        خطوات التنفيذ:
        1️⃣ معالجة أعمدة الأرقام
        2️⃣ معالجة أعمدة التاريخ
        3️⃣ فحص الترحيل العشوائي
        4️⃣ إنشاء سجل العمليات
        5️⃣ التصدير لـ Excel (اختياري)
        """
        logger.info("=" * 60)
        logger.info("🚀 بدء خط الإنتاج الشامل")
        logger.info("=" * 60)
        
        start_time = datetime.now()
        
        # المرحلة 1: معالجة الأعمدة الرقمية
        self.process_numeric_columns()
        
        # المرحلة 2: معالجة أعمدة التاريخ
        self.process_date_columns()
        
        # المرحلة 3: فحص الترحيل
        self.detect_data_shifting()
        
        # المرحلة 4: إنشاء التقرير
        audit_report = self.generate_audit_report()
        
        # المرحلة 5: التصدير
        if output_file:
            self.export_to_excel(output_file)
        
        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        
        logger.info("=" * 60)
        logger.info("✅ تم إكمال خط الإنتاج بنجاح")
        logger.info(f"⏱️  المدة: {duration:.2f} ثانية")
        logger.info(f"📊 الخلايا المفحوصة: {audit_report['summary']['cells_checked']:,}")
        logger.info(f"🔧 الخلايا المعدلة: {audit_report['summary']['cells_modified']:,}")
        logger.info(f"📝 الملاحظات المضافة: {audit_report['summary']['notes_added']}")
        logger.info(f"📈 نسبة الفحص: {audit_report['summary']['coverage_percentage']:.2f}%")
        logger.info("=" * 60)
        
        return {
            'dataframe': self.df,
            'audit_report': audit_report,
            'duration_seconds': duration,
            'status': 'success'
        }


# ============================================================
# مثال على الاستخدام (Usage Example)
# ============================================================

if __name__ == "__main__":
    # إنشاء بيانات نموذجية للاختبار
    sample_data = {
        'ID': ['1', '2.0', 'ABC', '4.0', '5'],
        'Name': ['أحمد', 'فاطمة', 'محمد', 'علي', 'سارة'],
        'Date': ['2024-01-15', '2024-02-20 14:30:00', 'بدون تاريخ', '2024-04-10 00:00:00', None],
        'Price': [100, 200.50, 300.0, 'غير متوفر', 500.0],
        'Created': ['2024-01-01', '2024-01-02 08:15:00', '2024-01-03', 'تاريخ غير معروف', '2024-01-05']
    }
    
    df = pd.DataFrame(sample_data)
    
    print("📥 البيانات الأصلية:")
    print(df)
    print()
    
    # إنشاء محرك المعالجة
    engine = ComprehensiveProcessingEngine(
        df=df,
        date_columns=['Date', 'Created'],
        numeric_columns=['ID', 'Price']
    )
    
    # تشغيل خط الإنتاج
    result = engine.run_complete_pipeline(output_file='processed_data.xlsx')
    
    print("\n📤 البيانات المعالجة:")
    print(result['dataframe'])
    print()
    
    print("📊 ملخص التقرير:")
    summary = result['audit_report']['summary']
    for key, value in summary.items():
        print(f"  • {key}: {value}")
